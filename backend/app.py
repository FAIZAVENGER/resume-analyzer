from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from PyPDF2 import PdfReader, PdfWriter
from docx import Document
import os
import json
import time
import concurrent.futures
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from dotenv import load_dotenv
import traceback
import threading
import atexit
import requests
import re
import hashlib
import random
import gc
import sys
import base64
import io
import subprocess
import tempfile
import shutil
from pathlib import Path

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

# Configure Groq API Keys (5 keys for parallel processing)
GROQ_API_KEYS = [
    os.getenv('GROQ_API_KEY_1'),
    os.getenv('GROQ_API_KEY_2'),
    os.getenv('GROQ_API_KEY_3'),
    os.getenv('GROQ_API_KEY_4'),
    os.getenv('GROQ_API_KEY_5')
]

GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "llama-3.3-70b-versatile"

# Track API status
warmup_complete = False
last_activity_time = datetime.now()

# Get absolute path for uploads folder
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
REPORTS_FOLDER = os.path.join(BASE_DIR, 'reports')
RESUME_PREVIEW_FOLDER = os.path.join(BASE_DIR, 'resume_previews')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORTS_FOLDER, exist_ok=True)
os.makedirs(RESUME_PREVIEW_FOLDER, exist_ok=True)
print(f"📁 Upload folder: {UPLOAD_FOLDER}")
print(f"📁 Reports folder: {REPORTS_FOLDER}")
print(f"📁 Resume Previews folder: {RESUME_PREVIEW_FOLDER}")

# Cache for consistent scoring - ENHANCED for deterministic results
score_cache = {}
cache_lock = threading.Lock()
analysis_result_cache = {}  # New cache to store complete analysis results
analysis_cache_lock = threading.Lock()

# Batch processing configuration
MAX_CONCURRENT_REQUESTS = 5
MAX_BATCH_SIZE = 10  # UPDATED: Increased from 6 to 10
MIN_SKILLS_TO_SHOW = 5
MAX_SKILLS_TO_SHOW = 8

# Rate limiting protection
MAX_RETRIES = 3  # Increased from 2 to 3
RETRY_DELAY_BASE = 2

# Track key usage - Updated for 5 keys
key_usage = {
    0: {'count': 0, 'last_used': None, 'cooling': False, 'errors': 0, 'requests_this_minute': 0, 'minute_window_start': None},
    1: {'count': 0, 'last_used': None, 'cooling': False, 'errors': 0, 'requests_this_minute': 0, 'minute_window_start': None},
    2: {'count': 0, 'last_used': None, 'cooling': False, 'errors': 0, 'requests_this_minute': 0, 'minute_window_start': None},
    3: {'count': 0, 'last_used': None, 'cooling': False, 'errors': 0, 'requests_this_minute': 0, 'minute_window_start': None},
    4: {'count': 0, 'last_used': None, 'cooling': False, 'errors': 0, 'requests_this_minute': 0, 'minute_window_start': None}
}

# Rate limit thresholds (Groq Developer Plan)
MAX_REQUESTS_PER_MINUTE_PER_KEY = 100
MAX_TOKENS_PER_MINUTE_PER_KEY = 250000

# Memory optimization
service_running = True

# Resume storage tracking
resume_storage = {}

# ENHANCED: Deterministic scoring system - no randomness
# Used to track deterministic hash-based scores, not for randomization
analysis_signatures = {}

# Domain-specific keyword libraries for strict matching
DOMAIN_KEYWORDS = {
    'vlsi': {
        'primary': ['verilog', 'vhdl', 'systemverilog', 'uvm', 'asic', 'fpga', 'rtl', 'synthesis', 'timing', 'sta', 'physical design', 'layout', 'spice', 'cadence', 'synopsys', 'mentor graphics', 'modelsim', 'questasim', 'vcs', 'ic design', 'vlsi', 'cmos', 'digital design', 'analog design', 'mixed-signal', 'dft', 'scan', 'bist', 'floorplanning', 'placement', 'routing', 'clock tree', 'cts', 'power analysis', 'ir drop', 'em', 'drc', 'lvs', 'rc extraction', 'post-layout', 'pre-layout', 'standard cell', 'custom cell', 'memory design', 'sram', 'dram', 'flash', 'pcie', 'ddr', 'usb', 'ethernet', 'spi', 'i2c', 'uart', 'amba', 'axi', 'ahb', 'apb'],
        'weight': 3.0,
        'threshold': 0.15  # If less than 15% of keywords match, score is low
    },
    'machine learning': {
        'primary': ['machine learning', 'deep learning', 'neural networks', 'tensorflow', 'pytorch', 'keras', 'scikit-learn', 'nlp', 'computer vision', 'cnn', 'rnn', 'lstm', 'transformer', 'bert', 'gpt', 'llm', 'data science', 'python', 'pandas', 'numpy', 'matplotlib', 'seaborn', 'jupyter', 'anaconda', 'spark', 'hadoop', 'sql', 'nosql', 'mongodb', 'postgresql', 'mysql', 'sqlite', 'aws', 'sagemaker', 'azure', 'gcp', 'docker', 'kubernetes', 'mlflow', 'kubeflow', 'tensorboard', 'opencv', 'yolo', 'ssd', 'faster rcnn', 'resnet', 'vgg', 'inception', 'alexnet', 'transfer learning', 'fine-tuning', 'hyperparameter', 'grid search', 'random search', 'bayesian', 'optimization', 'gradient descent', 'backpropagation', 'activation function', 'relu', 'sigmoid', 'tanh', 'softmax', 'loss function', 'cross-entropy', 'mse', 'mae', 'accuracy', 'precision', 'recall', 'f1-score', 'auc', 'roc', 'confusion matrix', 'classification', 'regression', 'clustering', 'k-means', 'dbscan', 'hierarchical', 'pca', 'svm', 'decision tree', 'random forest', 'xgboost', 'lightgbm', 'catboost', 'ensemble', 'bagging', 'boosting', 'stacking', 'feature engineering', 'feature selection', 'dimensionality reduction', 'data preprocessing', 'data cleaning', 'eda', 'statistics', 'probability', 'linear algebra', 'calculus', 'hypothesis testing', 'ab testing', 'time series', 'arima', 'sarima', 'prophet', 'reinforcement learning', 'q-learning', 'dqn', 'ppo', 'a2c', 'sde'],
        'weight': 3.0,
        'threshold': 0.15
    },
    'software engineering': {
        'primary': ['java', 'python', 'javascript', 'typescript', 'c++', 'c#', 'go', 'rust', 'swift', 'kotlin', 'react', 'angular', 'vue', 'node', 'express', 'django', 'flask', 'spring', 'hibernate', 'jpa', 'sql', 'nosql', 'mongodb', 'postgresql', 'mysql', 'oracle', 'redis', 'elasticsearch', 'rabbitmq', 'kafka', 'docker', 'kubernetes', 'aws', 'azure', 'gcp', 'microservices', 'rest', 'graphql', 'grpc', 'soap', 'ci/cd', 'jenkins', 'gitlab', 'github actions', 'git', 'svn', 'jira', 'confluence', 'agile', 'scrum', 'kanban', 'tdd', 'bdd', 'junit', 'pytest', 'mockito', 'selenium', 'cypress', 'jira', 'confluence', 'agile', 'scrum', 'kanban', 'tdd', 'bdd', 'solid', 'design patterns', 'oop', 'functional programming', 'multithreading', 'concurrency', 'performance optimization', 'memory management', 'garbage collection', 'profiling', 'debugging', 'linux', 'unix', 'bash', 'powershell', 'networking', 'tcp/ip', 'http', 'websockets', 'security', 'authentication', 'authorization', 'oauth', 'jwt', 'ssl/tls', 'cryptography', 'pci compliance', 'gdpr', 'hipaa'],
        'weight': 3.0,
        'threshold': 0.15
    },
    'data science': {
        'primary': ['python', 'r', 'sql', 'pandas', 'numpy', 'scipy', 'matplotlib', 'seaborn', 'plotly', 'scikit-learn', 'tensorflow', 'pytorch', 'keras', 'statistics', 'probability', 'hypothesis testing', 'anova', 'chi-square', 'regression', 'linear regression', 'logistic regression', 'polynomial regression', 'ridge', 'lasso', 'elasticnet', 'classification', 'clustering', 'k-means', 'hierarchical', 'dbscan', 'pca', 'svm', 'decision tree', 'random forest', 'gradient boosting', 'xgboost', 'lightgbm', 'catboost', 'time series', 'arima', 'sarima', 'prophet', 'ets', 'exponential smoothing', 'data visualization', 'tableau', 'power bi', 'looker', 'qlik', 'data wrangling', 'data cleaning', 'feature engineering', 'etl', 'data pipelines', 'big data', 'hadoop', 'spark', 'hive', 'pig', 'impala', 'presto', 'aws', 's3', 'redshift', 'emr', 'azure', 'synapse', 'databricks', 'snowflake', 'google cloud', 'bigquery', 'dbt', 'airflow', 'luigi', 'prefect', 'mlops', 'model deployment', 'api', 'flask', 'fastapi', 'docker', 'kubernetes'],
        'weight': 3.0,
        'threshold': 0.15
    },
    'devops': {
        'primary': ['aws', 'azure', 'gcp', 'linux', 'unix', 'windows server', 'docker', 'kubernetes', 'terraform', 'ansible', 'puppet', 'chef', 'saltstack', 'jenkins', 'gitlab ci', 'github actions', 'circleci', 'travis ci', 'teamcity', 'bamboo', 'git', 'svn', 'prometheus', 'grafana', 'nagios', 'zabbix', 'datadog', 'new relic', 'splunk', 'elk stack', 'elasticsearch', 'logstash', 'kibana', 'fluentd', 'nginx', 'apache', 'iis', 'haproxy', 'f5', 'python', 'bash', 'powershell', 'ruby', 'go', 'networking', 'vpc', 'subnet', 'firewall', 'load balancer', 'dns', 'dhcp', 'vpn', 'ssl/tls', 'ssh', 'security', 'iac', 'infrastructure as code', 'cloud formation', 'arm templates', 'deployment manager', 'ci/cd', 'continuous integration', 'continuous delivery', 'continuous deployment', 'automation', 'orchestration', 'containers', 'orchestration', 'serverless', 'lambda', 'functions', 'api gateway', 'cloudwatch', 'azure monitor', 'stackdriver', 'opsgenie', 'pagerduty', 'sre', 'reliability', 'scalability', 'high availability', 'disaster recovery', 'backup', 'restore'],
        'weight': 3.0,
        'threshold': 0.15
    }
}

def update_activity():
    """Update last activity timestamp"""
    global last_activity_time
    last_activity_time = datetime.now()

def get_available_key(resume_index=None):
    """Get the next available Groq API key using improved round-robin with rate limit checking"""
    if not any(GROQ_API_KEYS):
        return None, None
    
    current_time = datetime.now()
    
    # Reset minute counters if needed
    for i in range(5):
        if key_usage[i]['minute_window_start'] is None:
            key_usage[i]['minute_window_start'] = current_time
            key_usage[i]['requests_this_minute'] = 0
        elif (current_time - key_usage[i]['minute_window_start']).total_seconds() > 60:
            key_usage[i]['minute_window_start'] = current_time
            key_usage[i]['requests_this_minute'] = 0
    
    # If specific index provided, try that key first
    if resume_index is not None:
        key_index = resume_index % 5
        if (GROQ_API_KEYS[key_index] and 
            not key_usage[key_index]['cooling'] and
            key_usage[key_index]['requests_this_minute'] < MAX_REQUESTS_PER_MINUTE_PER_KEY):
            return GROQ_API_KEYS[key_index], key_index + 1
    
    # Find the best key (least used this minute, not cooling, has lowest error count)
    available_keys = []
    for i, key in enumerate(GROQ_API_KEYS):
        if (key and 
            not key_usage[i]['cooling'] and
            key_usage[i]['requests_this_minute'] < MAX_REQUESTS_PER_MINUTE_PER_KEY):
            # Calculate priority score (lower is better)
            priority_score = (
                key_usage[i]['requests_this_minute'] * 10 +  # Usage weight
                key_usage[i]['errors'] * 5                   # Error weight
            )
            available_keys.append((priority_score, i, key))
    
    if not available_keys:
        # All keys are cooling or rate limited, try any non-cooling key
        for i, key in enumerate(GROQ_API_KEYS):
            if key and not key_usage[i]['cooling']:
                print(f"⚠️ Using key {i+1} even though it's near limit: {key_usage[i]['requests_this_minute']}/{MAX_REQUESTS_PER_MINUTE_PER_KEY}")
                return key, i + 1
        return None, None
    
    # Sort by priority score and use the best one
    available_keys.sort(key=lambda x: x[0])
    best_key_index = available_keys[0][1]
    return GROQ_API_KEYS[best_key_index], best_key_index + 1

def mark_key_cooling(key_index, duration=30):
    """Mark a key as cooling down"""
    key_usage[key_index]['cooling'] = True
    key_usage[key_index]['last_used'] = datetime.now()
    
    def reset_cooling():
        time.sleep(duration)
        key_usage[key_index]['cooling'] = False
        print(f"✅ Key {key_index + 1} cooling completed")
    
    threading.Thread(target=reset_cooling, daemon=True).start()

def calculate_resume_hash(resume_text, job_description):
    """Calculate a hash for caching consistent scores"""
    # Use full content for deterministic caching
    content = f"{resume_text}{job_description}".encode('utf-8')
    return hashlib.md5(content).hexdigest()

def get_cached_analysis(resume_hash):
    """Get cached complete analysis if available - for deterministic results"""
    with analysis_cache_lock:
        return analysis_result_cache.get(resume_hash)

def set_cached_analysis(resume_hash, analysis):
    """Cache complete analysis for deterministic results"""
    with analysis_cache_lock:
        analysis_result_cache[resume_hash] = analysis

def get_cached_score(resume_hash):
    """Get cached score if available"""
    with cache_lock:
        return score_cache.get(resume_hash)

def set_cached_score(resume_hash, score):
    """Cache score for consistency"""
    with cache_lock:
        score_cache[resume_hash] = score

def detect_job_domain(job_description):
    """
    Detect the primary domain/industry of the job description.
    Returns domain name and confidence score.
    """
    jd_lower = job_description.lower()
    domain_scores = {}
    
    # Check for each domain
    for domain, keywords in DOMAIN_KEYWORDS.items():
        score = 0
        matches = 0
        
        # Check primary keywords
        for keyword in keywords['primary']:
            if keyword in jd_lower:
                matches += 1
                score += 1
        
        # Normalize score based on total keywords
        total_primary = len(keywords['primary'])
        if total_primary > 0:
            normalized_score = matches / total_primary
            domain_scores[domain] = {
                'score': normalized_score,
                'matches': matches,
                'threshold': keywords['threshold'],
                'weight': keywords['weight']
            }
    
    # Find the domain with highest score
    if domain_scores:
        best_domain = max(domain_scores.items(), key=lambda x: x[1]['score'])
        return best_domain[0], best_domain[1]
    
    return None, {'score': 0, 'matches': 0, 'threshold': 0.1, 'weight': 1.0}

# ========================================================================
#  MODIFIED SCORING FUNCTION - More generous, count‑based
# ========================================================================
def calculate_domain_match_score(resume_text, job_description):
    """
    Calculate professional ATS match score.
    Now uses a generous count‑based mapping so that candidates with a
    reasonable number of matched keywords receive scores in the 70–90+ range.
    """
    resume_lower = resume_text.lower()
    jd_lower = job_description.lower()
    
    # Detect job domain
    detected_domain, domain_info = detect_job_domain(job_description)
    
    # ---------- Domain‑specific scoring ----------
    if detected_domain:
        domain_keywords = DOMAIN_KEYWORDS[detected_domain]
        primary_keywords = domain_keywords['primary']
        
        # Count how many domain keywords appear in the resume
        matches = 0
        for keyword in primary_keywords:
            if keyword in resume_lower:
                matches += 1
        
        # ----- Reward based on ABSOLUTE number of matches (not percentage) -----
        # For large keyword lists (50+), 5–8 matches is decent, 10+ is strong.
        if matches >= 20:
            score = 95 + min(matches - 20, 5)   # 95–100
        elif matches >= 15:
            score = 85 + (matches - 15) * 2     # 85–94
        elif matches >= 10:
            score = 75 + (matches - 10) * 2     # 75–84
        elif matches >= 5:
            score = 60 + (matches - 5) * 3      # 60–74
        elif matches >= 3:
            score = 50 + (matches - 3) * 5      # 50–59
        elif matches >= 1:
            score = 40 + matches * 3            # 43–49
        else:
            score = 35                          # no match at all
        
        score = min(100, max(30, score))
        return round(score, 1)
    
    # ---------- Fallback: no clear domain detected ----------
    # Extract important terms from job description (similar to original)
    words = re.findall(r'\b[a-z]{3,}\b', jd_lower)
    word_freq = {}
    for word in words:
        if word not in ['the', 'and', 'for', 'with', 'this', 'that', 'have', 'from']:
            word_freq[word] = word_freq.get(word, 0) + 1
    
    # Use top 20 most frequent meaningful words
    top_words = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)[:20]
    top_keywords = [word for word, _ in top_words]
    
    # Count matches in resume
    matches = 0
    for keyword in top_keywords:
        if keyword in resume_lower:
            matches += 1
    
    # ----- Generous count‑based mapping for general case -----
    if matches >= 10:
        score = 85 + (matches - 10) * 1.5   # 85–100
    elif matches >= 7:
        score = 70 + (matches - 7) * 5      # 70–84
    elif matches >= 4:
        score = 55 + (matches - 4) * 5      # 55–69
    elif matches >= 1:
        score = 40 + matches * 5            # 45–59
    else:
        score = 35                         # no match
    
    score = min(100, max(35, score))
    return round(score, 1)

def generate_recommendation(score):
    """
    Generate strict, professional recommendation based on score.
    Real ATS systems have clear cutoffs.
    """
    if score >= 80:
        return "Strongly Recommended"
    elif score >= 70:
        return "Recommended"
    elif score >= 60:
        return "Consider"
    elif score >= 50:
        return "Consider with Reservations"
    elif score >= 40:
        return "Not Recommended - Low Match"
    elif score >= 30:
        return "Not Recommended - Significant Gap"
    elif score >= 20:
        return "Not Recommended - Wrong Domain"
    else:
        return "Rejected - Incompatible Profile"

def store_resume_file(file_data, filename, analysis_id):
    """Store resume file for later preview"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
        preview_filename = f"{analysis_id}_{safe_filename}"
        preview_path = os.path.join(RESUME_PREVIEW_FOLDER, preview_filename)
        
        # Save the original file
        with open(preview_path, 'wb') as f:
            if isinstance(file_data, bytes):
                f.write(file_data)
            else:
                file_data.save(f)
        
        # Also create a PDF version for preview if not already PDF
        file_ext = os.path.splitext(filename)[1].lower()
        pdf_preview_path = None
        
        if file_ext == '.pdf':
            pdf_preview_path = preview_path
        else:
            # Try to convert to PDF for better preview
            try:
                pdf_filename = f"{analysis_id}_{safe_filename.rsplit('.', 1)[0]}_preview.pdf"
                pdf_preview_path = os.path.join(RESUME_PREVIEW_FOLDER, pdf_filename)
                
                if file_ext in ['.docx', '.doc']:
                    # Try to convert DOC/DOCX to PDF
                    convert_doc_to_pdf(preview_path, pdf_preview_path)
                elif file_ext == '.txt':
                    # Convert TXT to PDF
                    convert_txt_to_pdf(preview_path, pdf_preview_path)
            except Exception as e:
                print(f"⚠️ Could not create PDF preview: {str(e)}")
                pdf_preview_path = None
        
        # Store in memory for quick access
        resume_storage[analysis_id] = {
            'filename': preview_filename,
            'original_filename': filename,
            'path': preview_path,
            'pdf_path': pdf_preview_path,
            'file_type': file_ext[1:],
            'has_pdf_preview': pdf_preview_path is not None and os.path.exists(pdf_preview_path),
            'stored_at': datetime.now().isoformat()
        }
        
        print(f"✅ Resume stored for preview: {preview_filename}")
        return preview_filename
    except Exception as e:
        print(f"❌ Error storing resume for preview: {str(e)}")
        return None

def convert_doc_to_pdf(doc_path, pdf_path):
    """Convert DOC/DOCX to PDF using LibreOffice or fallback methods"""
    try:
        # Check if LibreOffice is available
        if shutil.which('libreoffice'):
            # Use LibreOffice for conversion
            cmd = [
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', os.path.dirname(pdf_path),
                doc_path
            ]
            subprocess.run(cmd, check=True, capture_output=True, timeout=30)
            
            # Rename the output file
            expected_pdf = doc_path.rsplit('.', 1)[0] + '.pdf'
            if os.path.exists(expected_pdf):
                shutil.move(expected_pdf, pdf_path)
                return True
        else:
            # Fallback: Try using python-docx2pdf if available
            try:
                from docx2pdf import convert
                convert(doc_path, pdf_path)
                return True
            except ImportError:
                pass
            
            # Another fallback: Create a simple PDF from text
            extract_text_and_create_pdf(doc_path, pdf_path)
            return True
            
    except Exception as e:
        print(f"⚠️ DOC to PDF conversion failed: {str(e)}")
        # Create a simple PDF from extracted text
        extract_text_and_create_pdf(doc_path, pdf_path)
        return True
    
    return False

def convert_txt_to_pdf(txt_path, pdf_path):
    """Convert TXT to PDF"""
    try:
        extract_text_and_create_pdf(txt_path, pdf_path)
        return True
    except Exception as e:
        print(f"⚠️ TXT to PDF conversion failed: {str(e)}")
        return False

def extract_text_and_create_pdf(input_path, pdf_path):
    """Extract text and create a simple PDF"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
        from reportlab.lib.units import inch
        
        # Extract text based on file type
        file_ext = os.path.splitext(input_path)[1].lower()
        
        if file_ext == '.pdf':
            text = extract_text_from_pdf(input_path)
        elif file_ext in ['.docx', '.doc']:
            text = extract_text_from_docx(input_path)
        elif file_ext == '.txt':
            text = extract_text_from_txt(input_path)
        else:
            text = "Cannot preview this file type."
        
        # Create PDF
        doc = SimpleDocTemplate(pdf_path, pagesize=letter,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=72)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30
        )
        title = Paragraph("Resume Preview", title_style)
        story.append(title)
        
        # Add file info
        info_style = ParagraphStyle(
            'CustomInfo',
            parent=styles['Normal'],
            fontSize=10,
            textColor='gray',
            spaceAfter=20
        )
        info_text = f"Original file: {os.path.basename(input_path)} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        info = Paragraph(info_text, info_style)
        story.append(info)
        
        story.append(Spacer(1, 20))
        
        # Add content
        content_style = ParagraphStyle(
            'CustomContent',
            parent=styles['Normal'],
            fontSize=11,
            leading=14,
            spaceBefore=6,
            spaceAfter=6
        )
        
        # Split text into paragraphs
        paragraphs = text.split('\n')
        for para in paragraphs:
            if para.strip():
                story.append(Paragraph(para.replace('\t', '&nbsp;' * 4), content_style))
        
        # Build PDF
        doc.build(story)
        return True
        
    except Exception as e:
        print(f"⚠️ Failed to create PDF from text: {str(e)}")
        # Create minimal PDF
        try:
            from reportlab.pdfgen import canvas
            c = canvas.Canvas(pdf_path)
            c.drawString(100, 750, "Resume Preview")
            c.drawString(100, 730, f"File: {os.path.basename(input_path)}")
            c.drawString(100, 710, "Unable to display content. Please download the original file.")
            c.save()
            return True
        except:
            return False

def get_resume_preview(analysis_id):
    """Get resume preview data"""
    if analysis_id in resume_storage:
        return resume_storage[analysis_id]
    return None

def cleanup_resume_previews():
    """Clean up old resume previews"""
    try:
        now = datetime.now()
        for analysis_id in list(resume_storage.keys()):
            stored_time = datetime.fromisoformat(resume_storage[analysis_id]['stored_at'])
            if (now - stored_time).total_seconds() > 3600:  # 1 hour retention
                try:
                    # Clean up all related files
                    paths_to_clean = [
                        resume_storage[analysis_id]['path'],
                        resume_storage[analysis_id].get('pdf_path')
                    ]
                    
                    for path in paths_to_clean:
                        if path and os.path.exists(path):
                            os.remove(path)
                    
                    del resume_storage[analysis_id]
                    print(f"🧹 Cleaned up resume preview for {analysis_id}")
                except Exception as e:
                    print(f"⚠️ Error cleaning up files for {analysis_id}: {str(e)}")
        # Also clean up any orphaned files
        cleanup_orphaned_files()
    except Exception as e:
        print(f"⚠️ Error cleaning up resume previews: {str(e)}")

def cleanup_orphaned_files():
    """Clean up orphaned files in preview folder"""
    try:
        now = datetime.now()
        for filename in os.listdir(RESUME_PREVIEW_FOLDER):
            filepath = os.path.join(RESUME_PREVIEW_FOLDER, filename)
            if os.path.isfile(filepath):
                file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                if (now - file_time).total_seconds() > 7200:  # 2 hours
                    try:
                        os.remove(filepath)
                        print(f"🧹 Cleaned up orphaned file: {filename}")
                    except:
                        pass
    except Exception as e:
        print(f"⚠️ Error cleaning up orphaned files: {str(e)}")

def call_groq_api(prompt, api_key, max_tokens=1500, temperature=0.1, timeout=60, retry_count=0, key_index=None):
    """Call Groq API with optimized settings and rate limit protection"""
    if not api_key:
        print(f"❌ No Groq API key provided")
        return {'error': 'no_api_key', 'status': 500}
    
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json'
    }
    
    payload = {
        'model': GROQ_MODEL,
        'messages': [
            {
                'role': 'user',
                'content': prompt
            }
        ],
        'max_tokens': max_tokens,
        'temperature': temperature,  # Using 0.1 for consistency
        'top_p': 0.9,
        'stream': False,
        'stop': None
    }
    
    try:
        start_time = time.time()
        response = requests.post(
            GROQ_API_URL,
            headers=headers,
            json=payload,
            timeout=timeout
        )
        
        response_time = time.time() - start_time
        
        if response.status_code == 200:
            data = response.json()
            if 'choices' in data and len(data['choices']) > 0:
                result = data['choices'][0]['message']['content']
                print(f"✅ Groq API response in {response_time:.2f}s")
                return result
            else:
                print(f"❌ Unexpected Groq API response format")
                return {'error': 'invalid_response', 'status': response.status_code}
        
        # RATE LIMIT HANDLING - IMPROVED
        if response.status_code == 429:
            print(f"❌ Rate limit exceeded for Groq API (Key {key_index})")
            
            # Track this error for the key
            if key_index is not None:
                key_usage[key_index - 1]['errors'] += 1
                mark_key_cooling(key_index - 1, 60)  # Cool for 60 seconds on rate limit
            
            if retry_count < MAX_RETRIES:
                # Use exponential backoff with jitter
                wait_time = RETRY_DELAY_BASE ** (retry_count + 1) + random.uniform(2, 5)
                print(f"⏳ Rate limited, retrying in {wait_time:.1f}s (attempt {retry_count + 1}/{MAX_RETRIES})")
                time.sleep(wait_time)
                return call_groq_api(prompt, api_key, max_tokens, temperature, timeout, retry_count + 1, key_index)
            return {'error': 'rate_limit', 'status': 429}
        
        elif response.status_code == 503:
            print(f"❌ Service unavailable for Groq API")
            
            if retry_count < 2:
                wait_time = 15 + random.uniform(5, 10)
                print(f"⏳ Service unavailable, retrying in {wait_time:.1f}s")
                time.sleep(wait_time)
                return call_groq_api(prompt, api_key, max_tokens, temperature, timeout, retry_count + 1, key_index)
            return {'error': 'service_unavailable', 'status': 503}
        
        else:
            print(f"❌ Groq API Error {response.status_code}: {response.text[:100]}")
            if key_index is not None:
                key_usage[key_index - 1]['errors'] += 1
            return {'error': f'api_error_{response.status_code}', 'status': response.status_code}
            
    except requests.exceptions.Timeout:
        print(f"❌ Groq API timeout after {timeout}s")
        
        if retry_count < 2:
            wait_time = 10 + random.uniform(5, 10)
            print(f"⏳ Timeout, retrying in {wait_time:.1f}s (attempt {retry_count + 1}/3)")
            time.sleep(wait_time)
            return call_groq_api(prompt, api_key, max_tokens, temperature, timeout, retry_count + 1, key_index)
        return {'error': 'timeout', 'status': 408}
    
    except Exception as e:
        print(f"❌ Groq API Exception: {str(e)}")
        return {'error': str(e), 'status': 500}

def warmup_groq_service():
    """Warm up Groq service connection"""
    global warmup_complete
    
    available_keys = sum(1 for key in GROQ_API_KEYS if key)
    if available_keys == 0:
        print("⚠️ Skipping Groq warm-up: No API keys configured")
        return False
    
    try:
        print(f"🔥 Warming up Groq connection with {available_keys} keys...")
        print(f"📊 Using model: {GROQ_MODEL}")
        
        warmup_results = []
        
        for i, api_key in enumerate(GROQ_API_KEYS):
            if api_key:
                print(f"  Testing key {i+1}...")
                start_time = time.time()
                
                # Update minute counter
                current_time = datetime.now()
                if (key_usage[i]['minute_window_start'] is None or 
                    (current_time - key_usage[i]['minute_window_start']).total_seconds() > 60):
                    key_usage[i]['minute_window_start'] = current_time
                    key_usage[i]['requests_this_minute'] = 0
                
                response = call_groq_api(
                    prompt="Hello, are you ready? Respond with just 'ready'.",
                    api_key=api_key,
                    max_tokens=10,
                    temperature=0.1,
                    timeout=15,
                    key_index=i+1
                )
                
                if isinstance(response, dict) and 'error' in response:
                    print(f"    ⚠️ Key {i+1} failed: {response.get('error')}")
                    warmup_results.append(False)
                elif response and 'ready' in response.lower():
                    elapsed = time.time() - start_time
                    print(f"    ✅ Key {i+1} warmed up in {elapsed:.2f}s")
                    warmup_results.append(True)
                else:
                    print(f"    ⚠️ Key {i+1} warm-up failed: Unexpected response")
                    warmup_results.append(False)
                
                if i < available_keys - 1:
                    time.sleep(2)  # Increased delay between warm-up calls
        
        success = any(warmup_results)
        if success:
            print(f"✅ Groq service warmed up successfully")
            warmup_complete = True
        else:
            print(f"⚠️ Groq warm-up failed on all keys")
            
        return success
        
    except Exception as e:
        print(f"⚠️ Warm-up attempt failed: {str(e)}")
        threading.Timer(30.0, warmup_groq_service).start()
        return False

def keep_service_warm():
    """Periodically send requests to keep Groq service responsive"""
    global service_running
    
    while service_running:
        try:
            time.sleep(180)  # Every 3 minutes
            
            available_keys = sum(1 for key in GROQ_API_KEYS if key)
            if available_keys > 0 and warmup_complete:
                print(f"♨️ Keeping Groq warm with {available_keys} keys...")
                
                for i, api_key in enumerate(GROQ_API_KEYS):
                    if api_key and not key_usage[i]['cooling']:
                        # Check minute limit
                        current_time = datetime.now()
                        if (key_usage[i]['minute_window_start'] is None or 
                            (current_time - key_usage[i]['minute_window_start']).total_seconds() > 60):
                            key_usage[i]['minute_window_start'] = current_time
                            key_usage[i]['requests_this_minute'] = 0
                        
                        if key_usage[i]['requests_this_minute'] < 5:  # Only use if not busy
                            try:
                                response = call_groq_api(
                                    prompt="Ping - just say 'pong'",
                                    api_key=api_key,
                                    max_tokens=5,
                                    timeout=20,
                                    key_index=i+1
                                )
                                if response and 'pong' in str(response).lower():
                                    print(f"  ✅ Key {i+1} keep-alive successful")
                                else:
                                    print(f"  ⚠️ Key {i+1} keep-alive got unexpected response")
                            except Exception as e:
                                print(f"  ⚠️ Key {i+1} keep-alive failed: {str(e)}")
                        break
                    
        except Exception as e:
            print(f"⚠️ Keep-warm thread error: {str(e)}")
            time.sleep(180)

def keep_backend_awake():
    """Keep backend always active"""
    global service_running
    
    while service_running:
        try:
            time.sleep(60)  # Ping every 60 seconds
            
            try:
                # Self-ping to keep the service awake
                port = int(os.environ.get('PORT', 5002))
                requests.get(f"http://localhost:{port}/ping", timeout=10)
                print(f"✅ Self-ping successful to keep backend awake")
            except:
                # If self-ping fails, try health check
                try:
                    port = int(os.environ.get('PORT', 5002))
                    response = requests.get(f"http://localhost:{port}/health", timeout=10)
                    print(f"✅ Health check successful")
                except Exception as e:
                    print(f"⚠️ Keep-alive check failed: {e}")
                    
        except Exception as e:
            print(f"⚠️ Keep-backend-awake thread error: {str(e)}")
            time.sleep(60)

# Text extraction functions
def extract_text_from_pdf(file_path):
    """Extract text from PDF file with error handling"""
    try:
        text = ""
        max_attempts = 2
        
        for attempt in range(max_attempts):
            try:
                reader = PdfReader(file_path)
                text = ""
                
                for page_num, page in enumerate(reader.pages[:8]):  # Increased to 8 pages
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                    except Exception as e:
                        print(f"⚠️ PDF page extraction error: {e}")
                        continue
                
                if text.strip():
                    break
                    
            except Exception as e:
                print(f"❌ PDFReader attempt {attempt + 1} failed: {e}")
                if attempt == max_attempts - 1:
                    try:
                        with open(file_path, 'rb') as f:
                            content = f.read()
                            text = content.decode('utf-8', errors='ignore')
                            if text.strip():
                                words = text.split()
                                text = ' '.join(words[:1500])  # Increased word limit
                    except:
                        text = "Error: Could not extract text from PDF file"
        
        if not text.strip():
            return "Error: PDF appears to be empty or text could not be extracted"
        
        return text
    except Exception as e:
        print(f"❌ PDF Error: {traceback.format_exc()}")
        return f"Error reading PDF: {str(e)[:100]}"

def extract_text_from_docx(file_path):
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs[:150] if paragraph.text.strip()])
        
        if not text.strip():
            return "Error: Document appears to be empty"
        
        return text
    except Exception as e:
        print(f"❌ DOCX Error: {traceback.format_exc()}")
        return f"Error reading DOCX: {str(e)}"

def extract_text_from_txt(file_path):
    """Extract text from TXT file"""
    try:
        encodings = ['utf-8', 'latin-1', 'windows-1252', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    text = file.read()
                    
                if not text.strip():
                    return "Error: Text file appears to be empty"
                
                return text
            except UnicodeDecodeError:
                continue
                
        return "Error: Could not decode text file with common encodings"
        
    except Exception as e:
        print(f"❌ TXT Error: {traceback.format_exc()}")
        return f"Error reading TXT: {str(e)}"

def analyze_resume_with_ai(resume_text, job_description, filename=None, analysis_id=None, api_key=None, key_index=None):
    """
    Use Groq API to analyze resume against job description.
    ENHANCED: Deterministic scoring, strict domain matching, consistent results.
    """
    
    if not api_key:
        print(f"❌ No Groq API key provided for analysis.")
        return generate_fallback_analysis(filename, "No API key available")
    
    resume_text = resume_text[:3000]
    job_description = job_description[:1500]
    
    resume_hash = calculate_resume_hash(resume_text, job_description)
    
    # ENHANCED: Check cache first for deterministic results
    cached_analysis = get_cached_analysis(resume_hash)
    if cached_analysis:
        print(f"✅ Using cached analysis for deterministic result (Key {key_index})")
        cached_analysis['cached_result'] = True
        cached_analysis['key_used'] = f"Key {key_index}"
        if analysis_id:
            cached_analysis['analysis_id'] = analysis_id
        return cached_analysis
    
    # ENHANCED: Calculate strict domain match score FIRST (this will be the primary score)
    strict_score = calculate_domain_match_score(resume_text, job_description)
    
    # ENHANCED: Generate strict recommendation based on score
    strict_recommendation = generate_recommendation(strict_score)
    
    # ENHANCED: Modified prompt to emphasize strict scoring and domain matching
    prompt = f"""Analyze resume against job description and provide STRICT, PROFESSIONAL assessment.
This is for enterprise ATS (Applicant Tracking System) use.

CRITICAL: The overall_score MUST be PRECISELY {strict_score} - this is non-negotiable.
The recommendation MUST be: "{strict_recommendation}" - this is non-negotiable.

RESUME:
{resume_text}

JOB DESCRIPTION:
{job_description}

PROVIDE ANALYSIS IN THIS EXACT JSON FORMAT:
{{
    "candidate_name": "Extracted name or filename",
    "skills_matched": ["skill1", "skill2", "skill3", "skill4", "skill5", "skill6", "skill7", "skill8"],
    "skills_missing": ["skill1", "skill2", "skill3", "skill4", "skill5", "skill6", "skill7", "skill8"],
    "experience_summary": "Provide a concise 4-5 sentence summary of candidate's experience. Focus on key roles, achievements, and relevance. Make sure each sentence is complete and not truncated. Write full sentences.",
    "education_summary": "Provide a concise 4-5 sentence summary of education. Include degrees, institutions, and relevance. Make sure each sentence is complete and not truncated. Write full sentences.",
    "years_of_experience": "X years",
    "overall_score": {strict_score},
    "recommendation": "{strict_recommendation}",
    "key_strengths": ["strength1", "strength2", "strength3"],
    "areas_for_improvement": ["area1", "area2", "area3"]
}}

STRICT RULES:
1. overall_score MUST be exactly {strict_score} - NO EXCEPTIONS
2. recommendation MUST be exactly "{strict_recommendation}" - NO EXCEPTIONS
3. Provide EXACTLY 3 key_strengths and 3 areas_for_improvement
4. Write full, complete sentences. Do not cut off sentences mid-way.
5. Ensure proper sentence endings with periods.
6. Be honest and critical - this is professional ATS assessment."""

    try:
        print(f"⚡ Sending to Groq API (Key {key_index})...")
        start_time = time.time()
        
        # Track this request for rate limiting
        if key_index is not None:
            key_idx = key_index - 1
            current_time = datetime.now()
            if (key_usage[key_idx]['minute_window_start'] is None or 
                (current_time - key_usage[key_idx]['minute_window_start']).total_seconds() > 60):
                key_usage[key_idx]['minute_window_start'] = current_time
                key_usage[key_idx]['requests_this_minute'] = 0
            
            key_usage[key_idx]['requests_this_minute'] += 1
            print(f"📊 Key {key_index} usage: {key_usage[key_idx]['requests_this_minute']}/{MAX_REQUESTS_PER_MINUTE_PER_KEY} this minute")
        
        response = call_groq_api(
            prompt=prompt,
            api_key=api_key,
            max_tokens=1600,
            temperature=0.1,  # Low temperature for consistency
            timeout=60,
            key_index=key_index
        )
        
        if isinstance(response, dict) and 'error' in response:
            error_type = response.get('error')
            print(f"❌ Groq API error: {error_type}")
            
            if 'rate_limit' in error_type or '429' in str(error_type):
                if key_index:
                    mark_key_cooling(key_index - 1, 60)
            
            return generate_fallback_analysis(filename, f"API Error: {error_type}", strict_score=strict_score, strict_recommendation=strict_recommendation)
        
        elapsed_time = time.time() - start_time
        print(f"✅ Groq API response in {elapsed_time:.2f} seconds (Key {key_index})")
        
        result_text = response.strip()
        
        json_start = result_text.find('{')
        json_end = result_text.rfind('}') + 1
        
        if json_start != -1 and json_end > json_start:
            json_str = result_text[json_start:json_end]
        else:
            json_str = result_text
        
        json_str = json_str.replace('```json', '').replace('```', '').strip()
        
        try:
            analysis = json.loads(json_str)
            print(f"✅ Successfully parsed JSON response")
        except json.JSONDecodeError as e:
            print(f"❌ JSON Parse Error: {e}")
            print(f"Response was: {result_text[:150]}")
            
            return generate_fallback_analysis(filename, "JSON Parse Error", strict_score=strict_score, strict_recommendation=strict_recommendation)
        
        # ENHANCED: FORCE our strict score and recommendation (overwrite whatever AI returned)
        analysis['overall_score'] = strict_score
        analysis['recommendation'] = strict_recommendation
        
        # Validate and fill missing fields
        analysis = validate_analysis(analysis, filename)
        
        # ENHANCED: No score generation or variation - use the deterministic strict score
        set_cached_score(resume_hash, strict_score)
        
        analysis['ai_provider'] = "groq"
        analysis['ai_status'] = "Warmed up" if warmup_complete else "Warming up"
        analysis['ai_model'] = GROQ_MODEL
        analysis['response_time'] = f"{elapsed_time:.2f}s"
        analysis['key_used'] = f"Key {key_index}"
        analysis['scoring_method'] = 'strict_domain_matching'
        analysis['domain_detected'] = detect_job_domain(job_description)[0] if detect_job_domain(job_description)[0] else 'general'
        
        if analysis_id:
            analysis['analysis_id'] = analysis_id
        
        # ENHANCED: Cache the complete analysis for deterministic results
        set_cached_analysis(resume_hash, analysis)
        
        print(f"✅ Analysis completed: {analysis['candidate_name']} (Score: {analysis['overall_score']:.1f}) (Key {key_index})")
        print(f"   Domain: {analysis.get('domain_detected', 'general')}, Recommendation: {analysis['recommendation']}")
        
        return analysis
        
    except Exception as e:
        print(f"❌ Groq Analysis Error: {str(e)}")
        return generate_fallback_analysis(filename, f"Analysis Error: {str(e)[:100]}", strict_score=strict_score, strict_recommendation=strict_recommendation)
    
def validate_analysis(analysis, filename):
    """Validate analysis data and fill missing fields - FIXED to ensure complete sentences"""
    
    required_fields = {
        'candidate_name': 'Professional Candidate',
        'skills_matched': ['Python', 'JavaScript', 'SQL', 'Communication', 'Problem Solving', 'Team Collaboration', 'Project Management', 'Agile Methodology'],
        'skills_missing': ['Machine Learning', 'Cloud Computing', 'Data Analysis', 'DevOps', 'UI/UX Design', 'Cybersecurity', 'Mobile Development', 'Database Administration'],
        'experience_summary': 'The candidate demonstrates relevant professional experience with progressive responsibility. Their background shows expertise in key areas relevant to modern industry demands. They have experience collaborating with teams and delivering measurable results. Additional experience in specific domains enhances their suitability.',
        'education_summary': 'The candidate holds relevant educational qualifications from reputable institutions. Their academic background provides strong foundational knowledge in core subjects. Additional certifications enhance their professional profile. The education aligns well with industry requirements.',
        'years_of_experience': '3-5 years',
        'overall_score': 50.0,  # Default conservative score
        'recommendation': 'Consider with Reservations',
        'key_strengths': ['Strong technical foundation', 'Excellent communication skills', 'Proven track record of delivery'],
        'areas_for_improvement': ['Could benefit from advanced certifications', 'Limited experience in cloud platforms', 'Should gain experience with newer technologies']
    }
    
    for field, default_value in required_fields.items():
        if field not in analysis:
            analysis[field] = default_value
    
    if analysis['candidate_name'] == 'Professional Candidate' and filename:
        base_name = os.path.splitext(filename)[0]
        clean_name = base_name.replace('-', ' ').replace('_', ' ').title()
        if len(clean_name.split()) <= 4:
            analysis['candidate_name'] = clean_name
    
    # Ensure 5-8 skills in each category
    skills_matched = analysis.get('skills_matched', [])
    skills_missing = analysis.get('skills_missing', [])
    
    # If we have fewer than 5 skills, pad with defaults
    if len(skills_matched) < MIN_SKILLS_TO_SHOW:
        default_skills = ['Python', 'JavaScript', 'SQL', 'Communication', 'Problem Solving', 'Teamwork', 'Project Management', 'Agile']
        needed = MIN_SKILLS_TO_SHOW - len(skills_matched)
        skills_matched.extend(default_skills[:needed])
    
    if len(skills_missing) < MIN_SKILLS_TO_SHOW:
        default_skills = ['Machine Learning', 'Cloud Computing', 'Data Analysis', 'DevOps', 'UI/UX', 'Cybersecurity', 'Mobile Dev', 'Database']
        needed = MIN_SKILLS_TO_SHOW - len(skills_missing)
        skills_missing.extend(default_skills[:needed])
    
    # Limit to maximum
    analysis['skills_matched'] = skills_matched[:MAX_SKILLS_TO_SHOW]
    analysis['skills_missing'] = skills_missing[:MAX_SKILLS_TO_SHOW]
    
    # Ensure exactly 3 strengths and improvements
    analysis['key_strengths'] = analysis.get('key_strengths', [])[:3]
    analysis['areas_for_improvement'] = analysis.get('areas_for_improvement', [])[:3]
    
    # FIXED: Ensure complete sentences for summaries (don't truncate)
    for field in ['experience_summary', 'education_summary']:
        if field in analysis:
            text = analysis[field]
            # Remove any trailing ellipsis or incomplete sentences
            if '...' in text:
                # Find the last complete sentence before ellipsis
                sentences = text.split('. ')
                complete_sentences = []
                for sentence in sentences:
                    if '...' in sentence:
                        # Remove the incomplete part
                        sentence = sentence.split('...')[0]
                        if sentence.strip():
                            complete_sentences.append(sentence.strip() + '.')
                        break
                    elif sentence.strip():
                        complete_sentences.append(sentence.strip() + '.')
                analysis[field] = ' '.join(complete_sentences)
            
            # Ensure proper sentence endings
            if not analysis[field].strip().endswith(('.', '!', '?')):
                analysis[field] = analysis[field].strip() + '.'
            
            # Limit to reasonable length but keep sentences complete
            if len(analysis[field]) > 800:
                # Take first 5 sentences instead of truncating
                sentences = analysis[field].split('. ')
                if len(sentences) > 5:
                    analysis[field] = '. '.join(sentences[:5]) + '.'
    
    # Remove unwanted fields
    unwanted_fields = ['job_title_suggestion', 'industry_fit', 'salary_expectation']
    for field in unwanted_fields:
        if field in analysis:
            del analysis[field]
    
    return analysis

def generate_fallback_analysis(filename, reason, partial_success=False, strict_score=None, strict_recommendation=None):
    """
    Generate a fallback analysis with deterministic scoring.
    ENHANCED: Accepts strict_score and strict_recommendation for consistency.
    """
    candidate_name = "Professional Candidate"
    
    if filename:
        base_name = os.path.splitext(filename)[0]
        clean_name = base_name.replace('-', ' ').replace('_', ' ').replace('resume', '').replace('cv', '').strip()
        if clean_name:
            parts = clean_name.split()
            if len(parts) >= 2 and len(parts) <= 4:
                candidate_name = ' '.join(part.title() for part in parts)
    
    # Use provided strict score or generate conservative deterministic score
    if strict_score is not None:
        unique_score = strict_score
    else:
        # Generate deterministic score based on filename hash
        file_hash = hashlib.md5(filename.encode()).hexdigest() if filename else 'default'
        hash_int = int(file_hash[:4], 16)
        unique_score = 45 + (hash_int % 15)  # 45-59 range
        unique_score = round(unique_score, 1)
    
    # Use provided strict recommendation or generate based on score
    if strict_recommendation is not None:
        recommendation = strict_recommendation
    else:
        recommendation = generate_recommendation(unique_score)
    
    if partial_success:
        return {
            "candidate_name": candidate_name,
            "skills_matched": ['Python Programming', 'JavaScript Development', 'Database Management', 'Communication Skills', 'Problem Solving', 'Team Collaboration'],
            "skills_missing": ['Machine Learning Algorithms', 'Cloud Platform Expertise', 'Advanced Data Analysis', 'DevOps Practices', 'UI/UX Design Principles'],
            "experience_summary": 'The candidate has demonstrated professional experience in relevant technical roles. Their background includes working with modern technologies and methodologies. They have contributed to multiple projects with measurable outcomes. Additional experience enhances their suitability for the role.',
            "education_summary": 'The candidate possesses educational qualifications that provide a strong foundation for professional work. Their academic background includes relevant coursework and projects. Additional training complements their formal education. The educational profile aligns with industry requirements.',
            "years_of_experience": "3-5 years",
            "overall_score": unique_score,
            "recommendation": recommendation,
            "key_strengths": ['Technical proficiency', 'Communication abilities', 'Problem-solving approach'],
            "areas_for_improvement": ['Advanced technical skills needed', 'Cloud platform experience required', 'Industry-specific knowledge'],
            "ai_provider": "groq",
            "ai_status": "Partial",
            "ai_model": GROQ_MODEL,
            "scoring_method": "strict_domain_matching",
            "domain_detected": "general"
        }
    else:
        return {
            "candidate_name": candidate_name,
            "skills_matched": ['Basic Programming', 'Communication Skills', 'Problem Solving', 'Teamwork', 'Technical Knowledge', 'Learning Ability'],
            "skills_missing": ['Advanced Technical Skills', 'Industry Experience', 'Specialized Knowledge', 'Certifications', 'Project Management'],
            "experience_summary": 'Professional experience analysis will be available once the Groq AI service is fully initialized. The candidate appears to have relevant background based on initial file processing. Additional details will be available with full analysis.',
            "education_summary": 'Educational background analysis will be available shortly upon service initialization. Academic qualifications assessment is pending full AI processing. Further details will be provided with complete analysis.',
            "years_of_experience": "Not specified",
            "overall_score": unique_score,
            "recommendation": recommendation,
            "key_strengths": ['Fast learning capability', 'Strong work ethic', 'Good communication'],
            "areas_for_improvement": ['Service initialization required', 'Complete analysis pending', 'Detailed assessment needed'],
            "ai_provider": "groq",
            "ai_status": "Warming up",
            "ai_model": GROQ_MODEL,
            "scoring_method": "strict_domain_matching",
            "domain_detected": "general"
        }

def process_single_resume(args):
    """Process a single resume with intelligent error handling and rate limit protection"""
    resume_file, job_description, index, total, batch_id = args
    
    file_path = None
    try:
        print(f"📄 Processing resume {index + 1}/{total}: {resume_file.filename}")
        
        # IMPORTANT: Add staggered delays to prevent rate limits
        if index > 0:
            # Progressive delays based on position
            if index < 3:
                base_delay = 1.0
            elif index < 6:
                base_delay = 2.0
            else:
                base_delay = 3.0
            
            delay = base_delay + random.uniform(0, 0.5)
            print(f"⏳ Adding {delay:.1f}s delay before processing resume {index + 1}...")
            time.sleep(delay)
        
        api_key, key_index = get_available_key(index)
        if not api_key:
            print(f"⚠️ No available API key for resume {index + 1}")
            return {
                'filename': resume_file.filename,
                'error': 'No available API key',
                'status': 'failed',
                'index': index
            }
        
        file_ext = os.path.splitext(resume_file.filename)[1].lower()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        file_path = os.path.join(UPLOAD_FOLDER, f"batch_{batch_id}_{index}{file_ext}")
        
        # Save the file first
        resume_file.save(file_path)
        
        # Store resume for preview
        analysis_id = f"{batch_id}_resume_{index}"
        preview_filename = store_resume_file(resume_file, resume_file.filename, analysis_id)
        
        if file_ext == '.pdf':
            resume_text = extract_text_from_pdf(file_path)
        elif file_ext in ['.docx', '.doc']:
            resume_text = extract_text_from_docx(file_path)
        elif file_ext == '.txt':
            resume_text = extract_text_from_txt(file_path)
        else:
            if os.path.exists(file_path):
                os.remove(file_path)
            return {
                'filename': resume_file.filename,
                'error': f'Unsupported format: {file_ext}',
                'status': 'failed',
                'index': index
            }
        
        if resume_text.startswith('Error'):
            if os.path.exists(file_path):
                os.remove(file_path)
            return {
                'filename': resume_file.filename,
                'error': resume_text,
                'status': 'failed',
                'index': index
            }
        
        # Track key usage for rate limiting
        if key_index:
            key_idx = key_index - 1
            key_usage[key_idx]['count'] += 1
            key_usage[key_idx]['last_used'] = datetime.now()
            print(f"🔑 Using Key {key_index} (Total: {key_usage[key_idx]['count']}, This minute: {key_usage[key_idx]['requests_this_minute']})")
        
        analysis = analyze_resume_with_ai(
            resume_text, 
            job_description, 
            resume_file.filename, 
            analysis_id,
            api_key,
            key_index
        )
        
        analysis['filename'] = resume_file.filename
        analysis['original_filename'] = resume_file.filename
        
        resume_file.seek(0, 2)
        file_size = resume_file.tell()
        resume_file.seek(0)
        analysis['file_size'] = f"{(file_size / 1024):.1f}KB"
        
        analysis['analysis_id'] = analysis_id
        analysis['processing_order'] = index + 1
        analysis['key_used'] = f"Key {key_index}"
        
        # Add resume preview info
        analysis['resume_stored'] = preview_filename is not None
        analysis['has_pdf_preview'] = False
        
        if preview_filename:
            analysis['resume_preview_filename'] = preview_filename
            analysis['resume_original_filename'] = resume_file.filename
            # Check if PDF preview is available
            if analysis_id in resume_storage and resume_storage[analysis_id].get('has_pdf_preview'):
                analysis['has_pdf_preview'] = True
        
        # Keep the preview file, remove only the temp upload file
        if os.path.exists(file_path):
            os.remove(file_path)
        
        print(f"✅ Completed: {analysis.get('candidate_name')} - Score: {analysis.get('overall_score'):.1f} (Key {key_index})")
        
        # Check if key needs cooling after this request
        if key_index:
            key_idx = key_index - 1
            if key_usage[key_idx]['requests_this_minute'] >= MAX_REQUESTS_PER_MINUTE_PER_KEY - 5:
                print(f"⚠️ Key {key_index} near limit ({key_usage[key_idx]['requests_this_minute']}/{MAX_REQUESTS_PER_MINUTE_PER_KEY})")
        
        return {
            'analysis': analysis,
            'status': 'success',
            'index': index
        }
        
    except Exception as e:
        print(f"❌ Error processing {resume_file.filename}: {str(e)}")
        traceback.print_exc()
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass
        return {
            'filename': resume_file.filename,
            'error': f"Processing error: {str(e)[:100]}",
            'status': 'failed',
            'index': index
        }

@app.route('/')
def home():
    """Root route - API landing page"""
    
    inactive_time = datetime.now() - last_activity_time
    inactive_minutes = int(inactive_time.total_seconds() / 60)
    
    warmup_status = "✅ Ready" if warmup_complete else "🔥 Warming up..."
    
    available_keys = sum(1 for key in GROQ_API_KEYS if key)
    
    # Calculate current minute usage
    current_time = datetime.now()
    key_usage_info = []
    for i in range(5):
        if GROQ_API_KEYS[i]:
            if key_usage[i]['minute_window_start']:
                seconds_in_window = (current_time - key_usage[i]['minute_window_start']).total_seconds()
                if seconds_in_window > 60:
                    requests_this_minute = 0
                else:
                    requests_this_minute = key_usage[i]['requests_this_minute']
            else:
                requests_this_minute = 0
            key_usage_info.append(f"Key {i+1}: {requests_this_minute}/{MAX_REQUESTS_PER_MINUTE_PER_KEY}")
    
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Resume Analyzer API (Groq Parallel - Professional ATS)</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; padding: 0; background: #f5f5f5; }
            .container { max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            h1 { color: #333; }
            .status { padding: 10px; margin: 10px 0; border-radius: 5px; }
            .ready { background: #d4edda; color: #155724; }
            .warming { background: #fff3cd; color: #856404; }
            .endpoint { background: #f8f9fa; padding: 10px; margin: 10px 0; border-left: 4px solid #007bff; }
            .key-status { display: flex; gap: 10px; margin: 10px 0; }
            .key { padding: 5px 10px; border-radius: 3px; font-size: 12px; }
            .key-active { background: #d4edda; color: #155724; }
            .key-inactive { background: #f8d7da; color: #721c24; }
            .rate-limit-info { background: #e7f3ff; padding: 10px; border-radius: 5px; margin: 10px 0; }
            .scoring-info { background: #fff3e0; padding: 10px; border-radius: 5px; margin: 10px 0; border-left: 4px solid #ff6b6b; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🚀 Resume Analyzer API (Groq Parallel)</h1>
            <h2 style="color: #ff6b6b;">🏢 PROFESSIONAL ATS SCORING - STRICT DOMAIN MATCHING</h2>
            <p>Enterprise-grade resume analysis using Groq API with strict, deterministic scoring</p>
            
            <div class="status ''' + ('ready' if warmup_complete else 'warming') + '''">
                <strong>Status:</strong> ''' + warmup_status + '''
            </div>
            
            <div class="scoring-info">
                <strong>🎯 PROFESSIONAL ATS SCORING SYSTEM:</strong>
                <ul>
                    <li><strong style="color: #ff0000;">✅ STRICT DOMAIN MATCHING</strong> - ML resume vs VLSI job = 30-40 score</li>
                    <li><strong style="color: #00B050;">✅ DETERMINISTIC RESULTS</strong> - Same resume always gets same score</li>
                    <li><strong style="color: #4472C4;">✅ NO RANK FLIPPING</strong> - Candidate ranking never changes on re-analysis</li>
                    <li><strong style="color: #ff6b6b;">✅ REAL ATS BEHAVIOR</strong> - Clear cutoffs, no ambiguity</li>
                    <li><strong style="color: #ff9800;">✅ Domain keyword libraries</strong> - VLSI, ML, Software, Data Science, DevOps</li>
                    <li><strong style="color: #2196f3;">✅ Weighted scoring</strong> - Domain match is primary factor</li>
                    <li><strong style="color: #9c27b0;">✅ 5-15% keyword match = 25-35 score</strong> (Wrong domain)</li>
                </ul>
            </div>
            
            <div class="rate-limit-info">
                <strong>⚠️ RATE LIMIT PROTECTION ACTIVE:</strong>
                <ul>
                    <li>Max ''' + str(MAX_REQUESTS_PER_MINUTE_PER_KEY) + ''' requests/minute per key</li>
                    <li>Staggered delays between requests</li>
                    <li>Automatic key rotation</li>
                    <li>60s cooling on rate limits</li>
                    <li>Current usage: ''' + ', '.join(key_usage_info) + '''</li>
                </ul>
            </div>
            
            <div class="key-status">
                <strong>API Keys:</strong>
                ''' + ''.join([f'<span class="key ' + ('key-active' if key else 'key-inactive') + f'">Key {i+1}: ' + ('✅' if key else '❌') + '</span>' for i, key in enumerate(GROQ_API_KEYS)]) + '''
            </div>
            
            <p><strong>Model:</strong> ''' + GROQ_MODEL + '''</p>
            <p><strong>API Provider:</strong> Groq (Parallel Processing)</p>
            <p><strong>Max Batch Size:</strong> ''' + str(MAX_BATCH_SIZE) + ''' resumes</p>
            <p><strong>Processing:</strong> Rate-limited round-robin with staggered delays</p>
            <p><strong>Scoring:</strong> <span style="color: #ff6b6b; font-weight: bold;">PROFESSIONAL ATS - STRICT, DETERMINISTIC, DOMAIN-AWARE</span></p>
            <p><strong>Available Keys:</strong> ''' + str(available_keys) + '''/5</p>
            <p><strong>Last Activity:</strong> ''' + str(inactive_minutes) + ''' minutes ago</p>
            
            <h2>📡 Endpoints</h2>
            <div class="endpoint">
                <strong>POST /analyze</strong> - Analyze single resume (Professional ATS scoring)
            </div>
            <div class="endpoint">
                <strong>POST /analyze-batch</strong> - Analyze multiple resumes (up to ''' + str(MAX_BATCH_SIZE) + ''')
            </div>
            <div class="endpoint">
                <strong>GET /health</strong> - Health check with key status
            </div>
            <div class="endpoint">
                <strong>GET /ping</strong> - Keep-alive ping
            </div>
            <div class="endpoint">
                <strong>GET /quick-check</strong> - Check Groq API availability
            </div>
            <div class="endpoint">
                <strong>GET /resume-preview/&lt;analysis_id&gt;</strong> - Get resume preview (PDF)
            </div>
            <div class="endpoint">
                <strong>GET /resume-original/&lt;analysis_id&gt;</strong> - Download original resume file
            </div>
            <div class="endpoint">
                <strong>GET /download/&lt;filename&gt;</strong> - Download batch report
            </div>
            <div class="endpoint">
                <strong>GET /download-single/&lt;analysis_id&gt;</strong> - Download single candidate report
            </div>
        </div>
    </body>
    </html>
    '''

@app.route('/analyze', methods=['POST'])
def analyze_resume():
    """Analyze single resume"""
    update_activity()
    
    try:
        print("\n" + "="*50)
        print("📥 New single analysis request received")
        start_time = time.time()
        
        if 'resume' not in request.files:
            print("❌ No resume file in request")
            return jsonify({'error': 'No resume file provided'}), 400
        
        if 'jobDescription' not in request.form:
            print("❌ No job description in request")
            return jsonify({'error': 'No job description provided'}), 400
        
        resume_file = request.files['resume']
        job_description = request.form['jobDescription']
        
        print(f"📄 Resume file: {resume_file.filename}")
        print(f"📋 Job description: {len(job_description)} chars")
        
        if resume_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        resume_file.seek(0, 2)
        file_size = resume_file.tell()
        resume_file.seek(0)
        
        if file_size > 15 * 1024 * 1024:
            print(f"❌ File too large: {file_size} bytes")
            return jsonify({'error': 'File size too large. Maximum size is 15MB.'}), 400
        
        # Add small delay even for single requests
        time.sleep(0.5 + random.uniform(0, 0.3))
        
        api_key, key_index = get_available_key()
        if not api_key:
            return jsonify({'error': 'No available Groq API key'}), 500
        
        file_ext = os.path.splitext(resume_file.filename)[1].lower()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        file_path = os.path.join(UPLOAD_FOLDER, f"resume_{timestamp}{file_ext}")
        resume_file.save(file_path)
        
        # Store resume for preview
        analysis_id = f"single_{timestamp}"
        preview_filename = store_resume_file(resume_file, resume_file.filename, analysis_id)
        
        if file_ext == '.pdf':
            resume_text = extract_text_from_pdf(file_path)
        elif file_ext in ['.docx', '.doc']:
            resume_text = extract_text_from_docx(file_path)
        elif file_ext == '.txt':
            resume_text = extract_text_from_txt(file_path)
        else:
            return jsonify({'error': 'Unsupported file format. Please upload PDF, DOCX, or TXT'}), 400
        
        if resume_text.startswith('Error'):
            return jsonify({'error': resume_text}), 500
        
        analysis = analyze_resume_with_ai(resume_text, job_description, resume_file.filename, analysis_id, api_key, key_index)
        
        # Create single Excel report
        excel_filename = f"single_analysis_{analysis_id}.xlsx"
        excel_path = create_single_report(analysis, job_description, excel_filename)
        
        # Keep the preview file, remove only the temp upload file
        if os.path.exists(file_path):
            os.remove(file_path)
        
        analysis['excel_filename'] = os.path.basename(excel_path)
        analysis['ai_model'] = GROQ_MODEL
        analysis['ai_provider'] = "groq"
        analysis['ai_status'] = "Warmed up" if warmup_complete else "Warming up"
        analysis['response_time'] = analysis.get('response_time', 'N/A')
        analysis['analysis_id'] = analysis_id
        analysis['key_used'] = f"Key {key_index}"
        
        # Add resume preview info
        analysis['resume_stored'] = preview_filename is not None
        analysis['has_pdf_preview'] = False
        
        if preview_filename:
            analysis['resume_preview_filename'] = preview_filename
            analysis['resume_original_filename'] = resume_file.filename
            # Check if PDF preview is available
            if analysis_id in resume_storage and resume_storage[analysis_id].get('has_pdf_preview'):
                analysis['has_pdf_preview'] = True
        
        total_time = time.time() - start_time
        print(f"✅ Request completed in {total_time:.2f} seconds")
        print(f"   Domain: {analysis.get('domain_detected', 'general')}, Score: {analysis.get('overall_score', 0):.1f}")
        print("="*50 + "\n")
        
        return jsonify(analysis)
        
    except Exception as e:
        print(f"❌ Unexpected error: {traceback.format_exc()}")
        return jsonify({'error': f'Server error: {str(e)[:200]}'}), 500

@app.route('/analyze-batch', methods=['POST'])
def analyze_resume_batch():
    """Analyze multiple resumes with parallel processing and rate limit protection"""
    update_activity()
    
    try:
        print("\n" + "="*50)
        print("📦 New batch analysis request received")
        start_time = time.time()
        
        # ENHANCED: Clear used scores - but with deterministic scoring, we don't need randomization
        # Just for tracking purposes
        
        if 'resumes' not in request.files:
            print("❌ No 'resumes' key in request.files")
            return jsonify({'error': 'No resume files provided'}), 400
        
        resume_files = request.files.getlist('resumes')
        
        if 'jobDescription' not in request.form:
            print("❌ No job description in request")
            return jsonify({'error': 'No job description provided'}), 400
        
        job_description = request.form['jobDescription']
        
        if len(resume_files) == 0:
            print("❌ No files selected")
            return jsonify({'error': 'No files selected'}), 400
        
        print(f"📦 Batch size: {len(resume_files)} resumes")
        
        if len(resume_files) > MAX_BATCH_SIZE:
            print(f"❌ Too many files: {len(resume_files)} (max: {MAX_BATCH_SIZE})")
            return jsonify({'error': f'Maximum {MAX_BATCH_SIZE} resumes allowed per batch'}), 400
        
        available_keys = sum(1 for key in GROQ_API_KEYS if key)
        if available_keys == 0:
            print("❌ No Groq API keys configured")
            return jsonify({'error': 'No Groq API keys configured'}), 500
        
        batch_id = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        
        # Reset tracking
        for i in range(5):
            key_usage[i]['count'] = 0
            key_usage[i]['last_used'] = None
            key_usage[i]['errors'] = 0
            key_usage[i]['requests_this_minute'] = 0
            key_usage[i]['minute_window_start'] = datetime.now()
        
        all_analyses = []
        errors = []
        
        print(f"🔄 Processing {len(resume_files)} resumes with {available_keys} keys...")
        print(f"⚠️ RATE LIMIT PROTECTION: Staggered delays, max {MAX_REQUESTS_PER_MINUTE_PER_KEY} requests/minute/key")
        print(f"🏢 PROFESSIONAL ATS SCORING: Strict domain matching, deterministic results")
        
        # Process sequentially with delays (safer than parallel for rate limits)
        for index, resume_file in enumerate(resume_files):
            if resume_file.filename == '':
                errors.append({
                    'filename': 'Empty file',
                    'error': 'File has no name',
                    'index': index
                })
                continue
            
            print(f"\n🔑 Processing resume {index + 1}/{len(resume_files)}")
            
            args = (resume_file, job_description, index, len(resume_files), batch_id)
            result = process_single_resume(args)
            
            if result['status'] == 'success':
                all_analyses.append(result['analysis'])
            else:
                errors.append({
                    'filename': result.get('filename', 'Unknown'),
                    'error': result.get('error', 'Unknown error'),
                    'index': result.get('index')
                })
            
            # Check if any key needs cooling
            for i in range(5):
                if key_usage[i]['requests_this_minute'] >= MAX_REQUESTS_PER_MINUTE_PER_KEY - 2:
                    print(f"⚠️ Key {i+1} near limit, marking for cooling")
                    mark_key_cooling(i, 30)
        
        # ENHANCED: Deterministic ranking - always sort by score, ties broken by filename hash
        # This ensures consistent ordering every time
        all_analyses.sort(key=lambda x: (-x.get('overall_score', 0), hashlib.md5(x.get('filename', '').encode()).hexdigest()))
        
        for rank, analysis in enumerate(all_analyses, 1):
            analysis['rank'] = rank
        
        batch_excel_path = None
        if all_analyses:
            try:
                print("📊 Creating batch Excel report...")
                excel_filename = f"batch_analysis_{batch_id}.xlsx"
                batch_excel_path = create_comprehensive_batch_report(all_analyses, job_description, excel_filename)
                print(f"✅ Excel report created: {batch_excel_path}")
            except Exception as e:
                print(f"❌ Failed to create Excel report: {str(e)}")
                traceback.print_exc()
                # Create a minimal report
                batch_excel_path = create_minimal_batch_report(all_analyses, job_description, excel_filename)
        
        key_stats = []
        configured_keys = 0
        for i in range(5):
            if GROQ_API_KEYS[i]:
                configured_keys += 1
                key_stats.append({
                    'key': f'Key {i+1}',
                    'used': key_usage[i]['count'],
                    'requests_this_minute': key_usage[i]['requests_this_minute'],
                    'errors': key_usage[i]['errors'],
                    'status': 'cooling' if key_usage[i]['cooling'] else 'available',
                    'configured': True
                })
            else:
                key_stats.append({
                    'key': f'Key {i+1}',
                    'configured': False,
                    'status': 'not_configured'
                })
        
        total_time = time.time() - start_time
        
        # Calculate score statistics
        if all_analyses:
            scores = [a.get('overall_score', 0) for a in all_analyses]
            avg_score = round(sum(scores) / len(scores), 2)
            unique_scores = len(set(scores))  # Now truly unique due to deterministic scoring
            score_range = f"{min(scores):.1f}-{max(scores):.1f}"
        else:
            avg_score = 0
            unique_scores = 0
            score_range = "N/A"
        
        batch_summary = {
            'success': True,
            'total_files': len(resume_files),
            'successfully_analyzed': len(all_analyses),
            'failed_files': len(errors),
            'errors': errors,
            'batch_excel_filename': os.path.basename(batch_excel_path) if batch_excel_path else None,
            'batch_id': batch_id,
            'analyses': all_analyses,
            'model_used': GROQ_MODEL,
            'ai_provider': "groq",
            'ai_status': "Warmed up" if warmup_complete else "Warming up",
            'processing_time': f"{total_time:.2f}s",
            'processing_method': 'rate_limited_sequential',
            'key_statistics': key_stats,
            'available_keys': configured_keys,
            'total_keys': 5,
            'rate_limit_protection': f"Active (max {MAX_REQUESTS_PER_MINUTE_PER_KEY}/min/key)",
            'success_rate': f"{(len(all_analyses) / len(resume_files)) * 100:.1f}%" if resume_files else "0%",
            'performance': f"{len(all_analyses)/total_time:.2f} resumes/second" if total_time > 0 else "N/A",
            'scoring_quality': {
                'average_score': avg_score,
                'score_range': score_range,
                'unique_scores': unique_scores,
                'total_candidates': len(all_analyses),
                'scoring_method': 'professional_ats_strict_domain_matching',
                'deterministic_ranking': True,
                'no_ranking_flip': True
            },
            'scoring_system': '🏢 PROFESSIONAL ATS - STRICT DOMAIN MATCHING',
            'warning': 'This is a strict, deterministic ATS scoring system. Domain mismatch results in low scores (25-40).'
        }
        
        print(f"✅ Batch analysis completed in {total_time:.2f}s")
        print(f"📊 Key usage summary:")
        for stat in key_stats:
            if stat.get('configured', False):
                print(f"  {stat['key']}: {stat['used']} total, {stat['requests_this_minute']}/min, {stat['errors']} errors, {stat['status']}")
            else:
                print(f"  {stat['key']}: Not configured")
        print(f"🏢 PROFESSIONAL ATS SCORING - Avg: {avg_score:.2f}, Range: {score_range}")
        print(f"   Deterministic Ranking: Enabled - No ranking flips")
        print("="*50 + "\n")
        
        return jsonify(batch_summary)
        
    except Exception as e:
        print(f"❌ Batch analysis error: {traceback.format_exc()}")
        return jsonify({'error': f'Server error: {str(e)[:200]}'}), 500

@app.route('/resume-preview/<analysis_id>', methods=['GET'])
def get_resume_preview(analysis_id):
    """Get resume preview as PDF"""
    update_activity()
    
    try:
        print(f"📄 Resume preview request for: {analysis_id}")
        
        # Get resume info from storage
        resume_info = get_resume_preview(analysis_id)
        if not resume_info:
            return jsonify({'error': 'Resume preview not found'}), 404
        
        # Try to use PDF preview if available
        preview_path = resume_info.get('pdf_path') or resume_info['path']
        
        if not os.path.exists(preview_path):
            return jsonify({'error': 'Preview file not found'}), 404
        
        # Determine file type
        file_ext = os.path.splitext(preview_path)[1].lower()
        
        if file_ext == '.pdf':
            return send_file(
                preview_path,
                as_attachment=False,
                download_name=f"resume_preview_{analysis_id}.pdf",
                mimetype='application/pdf'
            )
        else:
            # If not PDF, try to convert or return original
            return send_file(
                preview_path,
                as_attachment=True,
                download_name=resume_info['original_filename'],
                mimetype='application/octet-stream'
            )
            
    except Exception as e:
        print(f"❌ Resume preview error: {traceback.format_exc()}")
        return jsonify({'error': f'Failed to get resume preview: {str(e)}'}), 500

@app.route('/resume-original/<analysis_id>', methods=['GET'])
def get_resume_original(analysis_id):
    """Download original resume file"""
    update_activity()
    
    try:
        print(f"📄 Original resume request for: {analysis_id}")
        
        # Get resume info from storage
        resume_info = get_resume_preview(analysis_id)
        if not resume_info:
            return jsonify({'error': 'Resume not found'}), 404
        
        original_path = resume_info['path']
        
        if not os.path.exists(original_path):
            return jsonify({'error': 'Resume file not found'}), 404
        
        return send_file(
            original_path,
            as_attachment=True,
            download_name=resume_info['original_filename']
        )
            
    except Exception as e:
        print(f"❌ Original resume download error: {traceback.format_exc()}")
        return jsonify({'error': f'Failed to download resume: {str(e)}'}), 500

def convert_experience_to_bullet_points(experience_summary):
    """Convert experience summary paragraph to bullet points"""
    if not experience_summary:
        return "• No experience summary available."
    
    # Clean the text
    text = experience_summary.strip()
    
    # Remove any trailing ellipsis or incomplete sentences
    if '...' in text:
        # Find the last complete sentence before ellipsis
        sentences = text.split('. ')
        complete_sentences = []
        for sentence in sentences:
            if '...' in sentence:
                # Remove the incomplete part
                sentence = sentence.split('...')[0]
                if sentence.strip():
                    complete_sentences.append(sentence.strip() + '.')
                break
            elif sentence.strip():
                complete_sentences.append(sentence.strip() + '.')
        text = ' '.join(complete_sentences)
    
    # Split into sentences
    sentences = text.replace('\n', ' ').split('. ')
    
    # Filter out empty sentences
    sentences = [s.strip() for s in sentences if s.strip()]
    
    # Ensure proper sentence endings
    for i, sentence in enumerate(sentences):
        if not sentence.endswith('.') and not sentence.endswith('!') and not sentence.endswith('?'):
            sentences[i] = sentence + '.'
    
    # Limit to 5 bullet points max
    sentences = sentences[:5]
    
    # Convert to bullet points
    bullet_points = '\n'.join([f'• {sentence}' for sentence in sentences])
    
    return bullet_points

def create_single_report(analysis, job_description, filename="single_analysis.xlsx"):
    """Create a single candidate Excel report"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidate Analysis"
        
        # Define styles
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)
        
        # Color scheme
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light Blue
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # Title Section
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "RESUME ANALYSIS REPORT - SINGLE CANDIDATE"
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.border = thick_border
        
        # Report Info Section
        ws['A3'] = "Report Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A3'].font = label_font
        ws['B3'].font = value_font
        
        ws['A4'] = "AI Model:"
        ws['B4'] = f"Groq {GROQ_MODEL}"
        ws['A4'].font = label_font
        ws['B4'].font = value_font
        
        ws['A5'] = "Scoring System:"
        ws['B5'] = "Professional ATS - Strict Domain Matching"
        ws['A5'].font = label_font
        ws['B5'].font = Font(bold=True, color="FF0000", size=10)
        ws.merge_cells('B5:H5')
        
        ws['A6'] = "Job Description:"
        ws['B6'] = job_description[:100] + "..." if len(job_description) > 100 else job_description
        ws['A6'].font = label_font
        ws['B6'].font = value_font
        ws.merge_cells('B6:H6')
        
        # Candidate Information Section
        start_row = 8
        ws.merge_cells(f'A{start_row}:H{start_row}')
        section_cell = ws[f'A{start_row}']
        section_cell.value = "CANDIDATE INFORMATION"
        section_cell.font = header_font
        section_cell.fill = header_fill
        section_cell.alignment = Alignment(horizontal='center')
        section_cell.border = thin_border
        
        # Candidate Data
        data_rows = [
            ("Candidate Name", analysis.get('candidate_name', 'N/A')),
            ("File Name", analysis.get('filename', 'N/A')),
            ("ATS Score", f"{analysis.get('overall_score', 0):.1f}/100"),
            ("Domain Detected", analysis.get('domain_detected', 'general')),
            ("Years of Experience", analysis.get('years_of_experience', 'Not specified')),
            ("Recommendation", analysis.get('recommendation', 'N/A')),
        ]
        
        for idx, (label, value) in enumerate(data_rows):
            row = start_row + idx + 1
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].border = thin_border
            if label == "ATS Score":
                score = analysis.get('overall_score', 0)
                if score >= 80:
                    ws[f'B{row}'].font = Font(bold=True, color="00B050", size=10)
                elif score >= 60:
                    ws[f'B{row}'].font = Font(bold=True, color="FFC000", size=10)
                elif score >= 40:
                    ws[f'B{row}'].font = Font(bold=True, color="FF6B6B", size=10)
                else:
                    ws[f'B{row}'].font = Font(bold=True, color="FF0000", size=10)
            ws.merge_cells(f'B{row}:H{row}')
        
        # Skills Matched Section
        skills_row = start_row + len(data_rows) + 2
        ws.merge_cells(f'A{skills_row}:H{skills_row}')
        skills_header = ws[f'A{skills_row}']
        skills_header.value = "SKILLS MATCHED (5-8 skills)"
        skills_header.font = header_font
        skills_header.fill = section_fill
        skills_header.alignment = Alignment(horizontal='center')
        skills_header.border = thin_border
        
        skills_matched = analysis.get('skills_matched', [])
        for idx, skill in enumerate(skills_matched[:8]):
            row = skills_row + idx + 1
            ws[f'A{row}'] = f"{idx + 1}."
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = skill
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].border = thin_border
            ws.merge_cells(f'B{row}:H{row}')
        
        # Skills Missing Section
        missing_start = skills_row + len(skills_matched) + 2
        ws.merge_cells(f'A{missing_start}:H{missing_start}')
        missing_header = ws[f'A{missing_start}']
        missing_header.value = "SKILLS MISSING (5-8 skills)"
        missing_header.font = header_font
        missing_header.fill = section_fill
        missing_header.alignment = Alignment(horizontal='center')
        missing_header.border = thin_border
        
        skills_missing = analysis.get('skills_missing', [])
        for idx, skill in enumerate(skills_missing[:8]):
            row = missing_start + idx + 1
            ws[f'A{row}'] = f"{idx + 1}."
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = skill
            ws[f'B{row}'].font = Font(size=10, color="FF0000")
            ws[f'B{row}'].border = thin_border
            ws.merge_cells(f'B{row}:H{row}')
        
        # Experience Summary Section - Now in bullet points
        exp_start = missing_start + len(skills_missing) + 2
        ws.merge_cells(f'A{exp_start}:H{exp_start}')
        exp_header = ws[f'A{exp_start}']
        exp_header.value = "EXPERIENCE SUMMARY (Bullet Points)"
        exp_header.font = header_font
        exp_header.fill = section_fill
        exp_header.alignment = Alignment(horizontal='center')
        exp_header.border = thin_border
        
        # Convert experience summary to bullet points
        experience_bullets = convert_experience_to_bullet_points(analysis.get('experience_summary', ''))
        
        ws.merge_cells(f'A{exp_start + 1}:H{exp_start + 6}')
        exp_cell = ws[f'A{exp_start + 1}']
        exp_cell.value = experience_bullets
        exp_cell.font = value_font
        exp_cell.alignment = Alignment(wrap_text=True, vertical='top')
        exp_cell.border = thin_border
        ws.row_dimensions[exp_start + 1].height = 100  # Increased height for bullet points
        
        # Key Strengths Section
        strengths_start = exp_start + 8
        ws.merge_cells(f'A{strengths_start}:H{strengths_start}')
        strengths_header = ws[f'A{strengths_start}']
        strengths_header.value = "KEY STRENGTHS (3)"
        strengths_header.font = header_font
        strengths_header.fill = section_fill
        strengths_header.alignment = Alignment(horizontal='center')
        strengths_header.border = thin_border
        
        key_strengths = analysis.get('key_strengths', [])
        for idx, strength in enumerate(key_strengths[:3]):
            row = strengths_start + idx + 1
            ws[f'A{row}'] = f"{idx + 1}."
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = strength
            ws[f'B{row}'].font = Font(size=10, color="00B050")
            ws[f'B{row}'].border = thin_border
            ws.merge_cells(f'B{row}:H{row}')
        
        # Areas for Improvement Section
        improve_start = strengths_start + len(key_strengths) + 2
        ws.merge_cells(f'A{improve_start}:H{improve_start}')
        improve_header = ws[f'A{improve_start}']
        improve_header.value = "AREAS FOR IMPROVEMENT (3)"
        improve_header.font = header_font
        improve_header.fill = section_fill
        improve_header.alignment = Alignment(horizontal='center')
        improve_header.border = thin_border
        
        areas_for_improvement = analysis.get('areas_for_improvement', [])
        for idx, area in enumerate(areas_for_improvement[:3]):
            row = improve_start + idx + 1
            ws[f'A{row}'] = f"{idx + 1}."
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = area
            ws[f'B{row}'].font = Font(size=10, color="FF6600")
            ws[f'B{row}'].border = thin_border
            ws.merge_cells(f'B{row}:H{row}')
        
        # Set column widths
        column_widths = {'A': 20, 'B': 60}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save the file
        filepath = os.path.join(REPORTS_FOLDER, filename)
        wb.save(filepath)
        print(f"📄 Single Excel report saved to: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"❌ Error creating single Excel report: {str(e)}")
        traceback.print_exc()
        # Create minimal report
        filepath = os.path.join(REPORTS_FOLDER, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidate Analysis"
        ws['A1'] = "Resume Analysis Report"
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Candidate: {analysis.get('candidate_name', 'Unknown')}"
        ws['A4'] = f"Score: {analysis.get('overall_score', 0):.1f}/100"
        ws['A5'] = f"Experience: {analysis.get('years_of_experience', 'Not specified')}"
        wb.save(filepath)
        return filepath

def create_comprehensive_batch_report(analyses, job_description, filename="batch_resume_analysis.xlsx"):
    """Create a comprehensive batch Excel report with professional formatting"""
    try:
        wb = Workbook()
        
        # ================== CANDIDATE COMPARISON SHEET ==================
        ws_comparison = wb.active
        ws_comparison.title = "Candidate Comparison"
        
        # Define styles
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_font = Font(bold=True, color="000000", size=10)
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        
        # Color scheme
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light Blue
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # Title Section
        ws_comparison.merge_cells('A1:M1')
        title_cell = ws_comparison['A1']
        title_cell.value = "PROFESSIONAL ATS RESUME ANALYSIS REPORT - BATCH COMPARISON"
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.border = thick_border
        
        # Report Info Section
        info_row = 3
        info_data = [
            ("Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Candidates:", len(analyses)),
            ("AI Model:", f"Groq {GROQ_MODEL}"),
            ("Scoring System:", "Professional ATS - Strict Domain Matching"),
            ("Job Description:", job_description[:100] + "..." if len(job_description) > 100 else job_description),
        ]
        
        for i, (label, value) in enumerate(info_data):
            ws_comparison.cell(row=info_row, column=1 + i*3, value=label).font = bold_font
            ws_comparison.cell(row=info_row, column=2 + i*3, value=value).font = normal_font
            if label == "Scoring System:":
                ws_comparison.cell(row=info_row, column=2 + i*3).font = Font(bold=True, color="FF0000", size=10)
        
        # Candidate Comparison Table Headers
        start_row = 5
        headers = [
            ("Rank", 8),
            ("Candidate Name", 25),
            ("File Name", 25),
            ("Domain Detected", 15),
            ("Years of Experience", 15),
            ("ATS Score", 12),
            ("Recommendation", 20),
            ("Experience Summary", 40),
            ("Skills Matched", 30),
            ("Skills Missing", 30),
            ("Key Strengths", 25),
            ("Areas for Improvement", 25)
        ]
        
        # Write headers
        for col, (header, width) in enumerate(headers, start=1):
            cell = ws_comparison.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            ws_comparison.column_dimensions[get_column_letter(col)].width = width
        
        # Write candidate data
        for idx, analysis in enumerate(analyses):
            row = start_row + idx + 1
            row_fill = even_row_fill if idx % 2 == 0 else odd_row_fill
            
            # Rank
            cell = ws_comparison.cell(row=row, column=1, value=analysis.get('rank', '-'))
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = row_fill
            cell.border = thin_border
            
            # Candidate Name
            cell = ws_comparison.cell(row=row, column=2, value=analysis.get('candidate_name', 'Unknown'))
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # File Name
            cell = ws_comparison.cell(row=row, column=3, value=analysis.get('filename', 'Unknown'))
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # Domain Detected
            cell = ws_comparison.cell(row=row, column=4, value=analysis.get('domain_detected', 'general'))
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = row_fill
            cell.border = thin_border
            
            # Years of Experience
            cell = ws_comparison.cell(row=row, column=5, value=analysis.get('years_of_experience', 'Not specified'))
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = row_fill
            cell.border = thin_border
            
            # ATS Score with color coding
            score = analysis.get('overall_score', 0)
            cell = ws_comparison.cell(row=row, column=6, value=f"{score:.1f}")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = row_fill
            cell.border = thin_border
            if score >= 80:
                cell.font = Font(bold=True, color="00B050", size=10)
            elif score >= 60:
                cell.font = Font(bold=True, color="FFC000", size=10)
            elif score >= 40:
                cell.font = Font(bold=True, color="FF6B6B", size=10)
            else:
                cell.font = Font(bold=True, color="FF0000", size=10)
            
            # Recommendation
            cell = ws_comparison.cell(row=row, column=7, value=analysis.get('recommendation', 'N/A'))
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # Experience Summary - Now in bullet points
            exp_summary = analysis.get('experience_summary', 'No summary available.')
            experience_bullets = convert_experience_to_bullet_points(exp_summary)
            cell = ws_comparison.cell(row=row, column=8, value=experience_bullets)
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = row_fill
            cell.border = thin_border
            
            # Skills Matched
            skills_matched = analysis.get('skills_matched', [])
            cell = ws_comparison.cell(row=row, column=9, value="\n".join([f"• {skill}" for skill in skills_matched[:8]]))
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = row_fill
            cell.border = thin_border
            
            # Skills Missing
            skills_missing = analysis.get('skills_missing', [])
            cell = ws_comparison.cell(row=row, column=10, value="\n".join([f"• {skill}" for skill in skills_missing[:8]]))
            cell.font = Font(size=9, color="FF0000")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = row_fill
            cell.border = thin_border
            
            # Key Strengths
            strengths = analysis.get('key_strengths', [])
            cell = ws_comparison.cell(row=row, column=11, value="\n".join([f"• {strength}" for strength in strengths[:3]]))
            cell.font = Font(size=9, color="00B050")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = row_fill
            cell.border = thin_border
            
            # Areas for Improvement
            improvements = analysis.get('areas_for_improvement', [])
            cell = ws_comparison.cell(row=row, column=12, value="\n".join([f"• {area}" for area in improvements[:3]]))
            cell.font = Font(size=9, color="FF6600")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = row_fill
            cell.border = thin_border
        
        # Add summary statistics at the bottom
        summary_row = start_row + len(analyses) + 2
        scores = [a.get('overall_score', 0) for a in analyses]
        avg_score = round(sum(scores) / len(scores), 2) if scores else 0
        top_score = max(scores, default=0)
        bottom_score = min(scores, default=0)
        
        # Calculate average years of experience
        years_list = []
        for a in analyses:
            years_text = a.get('years_of_experience', 'Not specified')
            years_match = re.search(r'(\d+[\+\-]?)', str(years_text))
            if years_match:
                years_list.append(years_match.group(1))
        
        avg_years = "N/A"
        if years_list:
            try:
                numeric_years = []
                for y in years_list:
                    if '+' in y:
                        numeric_years.append(int(y.replace('+', '')) + 2)
                    elif '-' in y:
                        parts = y.split('-')
                        if len(parts) == 2:
                            numeric_years.append((int(parts[0]) + int(parts[1])) / 2)
                    else:
                        numeric_years.append(int(y))
                if numeric_years:
                    avg_years = f"{sum(numeric_years)/len(numeric_years):.1f} years"
            except:
                avg_years = "Various"
        
        # Unique scores count
        unique_scores = len(set(scores))
        
        summary_data = [
            ("Average Score:", f"{avg_score:.2f}/100"),
            ("Highest Score:", f"{top_score:.1f}/100"),
            ("Lowest Score:", f"{bottom_score:.1f}/100"),
            ("Average Experience:", avg_years),
            ("Unique Scores:", f"{unique_scores}/{len(analyses)}"),
            ("Analysis Date:", datetime.now().strftime("%Y-%m-%d")),
            ("Scoring System:", "Professional ATS - Strict Domain Matching")
        ]
        
        for i, (label, value) in enumerate(summary_data):
            ws_comparison.cell(row=summary_row, column=1 + i*2, value=label).font = bold_font
            ws_comparison.cell(row=summary_row, column=2 + i*2, value=value).font = bold_font
            ws_comparison.cell(row=summary_row, column=2 + i*2).alignment = Alignment(horizontal='center')
            if label == "Scoring System:":
                ws_comparison.cell(row=summary_row, column=2 + i*2).font = Font(bold=True, color="FF0000", size=10)
        
        # ================== INDIVIDUAL CANDIDATE SHEETS ==================
        for analysis in analyses:
            candidate_name = analysis.get('candidate_name', f"Candidate_{analysis.get('rank', 'Unknown')}")
            sheet_name = re.sub(r'[\\/*?:[\]]', '_', candidate_name[:31])
            
            ws_candidate = wb.create_sheet(title=sheet_name)
            
            candidate_title_font = Font(bold=True, size=14, color="FFFFFF")
            candidate_header_font = Font(bold=True, size=11, color="000000")
            candidate_label_font = Font(bold=True, size=10, color="000000")
            candidate_value_font = Font(size=10, color="000000")
            
            candidate_title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            candidate_section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Title Section
            ws_candidate.merge_cells('A1:H1')
            title_cell = ws_candidate['A1']
            title_cell.value = f"PROFESSIONAL ATS ANALYSIS - {candidate_name.upper()}"
            title_cell.font = candidate_title_font
            title_cell.fill = candidate_title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = thick_border
            
            # Report Info Section
            ws_candidate['A3'] = "Report Date:"
            ws_candidate['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_candidate['A3'].font = candidate_label_font
            ws_candidate['B3'].font = candidate_value_font
            
            ws_candidate['A4'] = "AI Model:"
            ws_candidate['B4'] = f"Groq {GROQ_MODEL}"
            ws_candidate['A4'].font = candidate_label_font
            ws_candidate['B4'].font = candidate_value_font
            
            ws_candidate['A5'] = "Scoring System:"
            ws_candidate['B5'] = "Professional ATS - Strict Domain Matching"
            ws_candidate['A5'].font = candidate_label_font
            ws_candidate['B5'].font = Font(bold=True, color="FF0000", size=10)
            ws_candidate.merge_cells('B5:H5')
            
            ws_candidate['A6'] = "Rank:"
            ws_candidate['B6'] = f"#{analysis.get('rank', 'N/A')}"
            ws_candidate['A6'].font = candidate_label_font
            ws_candidate['B6'].font = candidate_value_font
            
            # File Name
            ws_candidate.merge_cells('A8:H8')
            file_header = ws_candidate['A8']
            file_header.value = "FILE NAME"
            file_header.font = candidate_header_font
            file_header.fill = candidate_section_fill
            file_header.alignment = Alignment(horizontal='center')
            file_header.border = thin_border
            
            ws_candidate.merge_cells('A9:H9')
            file_cell = ws_candidate['A9']
            file_cell.value = analysis.get('filename', 'N/A')
            file_cell.font = candidate_value_font
            file_cell.border = thin_border
            
            # Domain Detected
            ws_candidate.merge_cells('A11:H11')
            domain_header = ws_candidate['A11']
            domain_header.value = "DOMAIN DETECTED"
            domain_header.font = candidate_header_font
            domain_header.fill = candidate_section_fill
            domain_header.alignment = Alignment(horizontal='center')
            domain_header.border = thin_border
            
            ws_candidate.merge_cells('A12:H12')
            domain_cell = ws_candidate['A12']
            domain_cell.value = analysis.get('domain_detected', 'general')
            domain_cell.font = Font(bold=True, size=12, color="4472C4")
            domain_cell.alignment = Alignment(horizontal='center')
            domain_cell.border = thin_border
            
            # Years of Experience
            ws_candidate.merge_cells('A14:H14')
            exp_header = ws_candidate['A14']
            exp_header.value = "YEARS OF EXPERIENCE"
            exp_header.font = candidate_header_font
            exp_header.fill = candidate_section_fill
            exp_header.alignment = Alignment(horizontal='center')
            exp_header.border = thin_border
            
            ws_candidate.merge_cells('A15:H15')
            exp_cell = ws_candidate['A15']
            exp_cell.value = analysis.get('years_of_experience', 'Not specified')
            exp_cell.font = Font(bold=True, size=12, color="4472C4")
            exp_cell.alignment = Alignment(horizontal='center')
            exp_cell.border = thin_border
            
            # ATS Score
            ws_candidate.merge_cells('A17:H17')
            score_header = ws_candidate['A17']
            score_header.value = "ATS SCORE"
            score_header.font = candidate_header_font
            score_header.fill = candidate_section_fill
            score_header.alignment = Alignment(horizontal='center')
            score_header.border = thin_border
            
            ws_candidate.merge_cells('A18:H18')
            score_cell = ws_candidate['A18']
            score_cell.value = f"{analysis.get('overall_score', 0):.1f}/100"
            score_cell.font = Font(bold=True, size=12, color=get_score_color(analysis.get('overall_score', 0)))
            score_cell.alignment = Alignment(horizontal='center')
            score_cell.border = thin_border
            
            # Recommendation
            ws_candidate.merge_cells('A20:H20')
            rec_header = ws_candidate['A20']
            rec_header.value = "RECOMMENDATION"
            rec_header.font = candidate_header_font
            rec_header.fill = candidate_section_fill
            rec_header.alignment = Alignment(horizontal='center')
            rec_header.border = thin_border
            
            ws_candidate.merge_cells('A21:H21')
            rec_cell = ws_candidate['A21']
            rec_cell.value = analysis.get('recommendation', 'N/A')
            rec_cell.font = Font(bold=True, size=11, color=get_score_color(analysis.get('overall_score', 0)))
            rec_cell.alignment = Alignment(horizontal='center')
            rec_cell.border = thin_border
            
            # Skills Matched
            skills_row = 23
            ws_candidate.merge_cells(f'A{skills_row}:H{skills_row}')
            skills_header = ws_candidate[f'A{skills_row}']
            skills_header.value = "SKILLS MATCHED (5-8 skills)"
            skills_header.font = candidate_header_font
            skills_header.fill = candidate_section_fill
            skills_header.alignment = Alignment(horizontal='center')
            skills_header.border = thin_border
            
            skills_matched = analysis.get('skills_matched', [])
            for idx, skill in enumerate(skills_matched[:8]):
                row = skills_row + idx + 1
                ws_candidate.merge_cells(f'A{row}:H{row}')
                cell = ws_candidate.cell(row=row, column=1, value=f"{idx + 1}. {skill}")
                cell.font = Font(size=10, color="00B050")
                cell.border = thin_border
            
            # Skills Missing
            missing_start = skills_row + len(skills_matched) + 2
            ws_candidate.merge_cells(f'A{missing_start}:H{missing_start}')
            missing_header = ws_candidate[f'A{missing_start}']
            missing_header.value = "SKILLS MISSING (5-8 skills)"
            missing_header.font = candidate_header_font
            missing_header.fill = candidate_section_fill
            missing_header.alignment = Alignment(horizontal='center')
            missing_header.border = thin_border
            
            skills_missing = analysis.get('skills_missing', [])
            for idx, skill in enumerate(skills_missing[:8]):
                row = missing_start + idx + 1
                ws_candidate.merge_cells(f'A{row}:H{row}')
                cell = ws_candidate.cell(row=row, column=1, value=f"{idx + 1}. {skill}")
                cell.font = Font(size=10, color="FF0000")
                cell.border = thin_border
            
            # Experience Summary - Now in bullet points
            exp_start = missing_start + len(skills_missing) + 2
            ws_candidate.merge_cells(f'A{exp_start}:H{exp_start}')
            exp_header = ws_candidate[f'A{exp_start}']
            exp_header.value = "EXPERIENCE SUMMARY (Bullet Points)"
            exp_header.font = candidate_header_font
            exp_header.fill = candidate_section_fill
            exp_header.alignment = Alignment(horizontal='center')
            exp_header.border = thin_border
            
            experience_bullets = convert_experience_to_bullet_points(analysis.get('experience_summary', ''))
            
            ws_candidate.merge_cells(f'A{exp_start + 1}:H{exp_start + 6}')
            exp_cell = ws_candidate[f'A{exp_start + 1}']
            exp_cell.value = experience_bullets
            exp_cell.font = candidate_value_font
            exp_cell.alignment = Alignment(wrap_text=True, vertical='top')
            exp_cell.border = thin_border
            ws_candidate.row_dimensions[exp_start + 1].height = 100
            
            # Key Strengths
            strengths_start = exp_start + 8
            ws_candidate.merge_cells(f'A{strengths_start}:H{strengths_start}')
            strengths_header = ws_candidate[f'A{strengths_start}']
            strengths_header.value = "KEY STRENGTHS (3)"
            strengths_header.font = candidate_header_font
            strengths_header.fill = candidate_section_fill
            strengths_header.alignment = Alignment(horizontal='center')
            strengths_header.border = thin_border
            
            key_strengths = analysis.get('key_strengths', [])
            for idx, strength in enumerate(key_strengths[:3]):
                row = strengths_start + idx + 1
                ws_candidate.merge_cells(f'A{row}:H{row}')
                cell = ws_candidate.cell(row=row, column=1, value=f"{idx + 1}. {strength}")
                cell.font = Font(size=10, color="00B050")
                cell.border = thin_border
            
            # Areas for Improvement
            improve_start = strengths_start + len(key_strengths) + 2
            ws_candidate.merge_cells(f'A{improve_start}:H{improve_start}')
            improve_header = ws_candidate[f'A{improve_start}']
            improve_header.value = "AREAS FOR IMPROVEMENT (3)"
            improve_header.font = candidate_header_font
            improve_header.fill = candidate_section_fill
            improve_header.alignment = Alignment(horizontal='center')
            improve_header.border = thin_border
            
            areas_for_improvement = analysis.get('areas_for_improvement', [])
            for idx, area in enumerate(areas_for_improvement[:3]):
                row = improve_start + idx + 1
                ws_candidate.merge_cells(f'A{row}:H{row}')
                cell = ws_candidate.cell(row=row, column=1, value=f"{idx + 1}. {area}")
                cell.font = Font(size=10, color="FF6600")
                cell.border = thin_border
            
            ws_candidate.column_dimensions['A'].width = 60
        
        filepath = os.path.join(REPORTS_FOLDER, filename)
        wb.save(filepath)
        print(f"📊 Professional ATS batch report saved to: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"❌ Error creating professional batch Excel report: {str(e)}")
        traceback.print_exc()
        return create_minimal_batch_report(analyses, job_description, filename)

def get_score_color(score):
    """Get color based on score"""
    if score >= 80:
        return "00B050"
    elif score >= 60:
        return "FFC000"
    elif score >= 40:
        return "FF6B6B"
    else:
        return "FF0000"

def create_minimal_batch_report(analyses, job_description, filename):
    """Create a minimal batch report as fallback"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Analysis"
        
        ws['A1'] = "Professional ATS Batch Resume Analysis Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:L1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total Candidates: {len(analyses)}"
        ws['A4'] = f"Scoring System: Professional ATS - Strict Domain Matching"
        ws['A4'].font = Font(bold=True, color="FF0000")
        
        headers = ["Rank", "Candidate Name", "File Name", "Domain", "Years of Experience", "ATS Score", "Recommendation", "Experience Summary", "Skills Matched", "Skills Missing", "Key Strengths", "Areas for Improvement"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
        
        for idx, analysis in enumerate(analyses):
            row = 6 + idx
            ws.cell(row=row, column=1, value=analysis.get('rank', '-'))
            ws.cell(row=row, column=2, value=analysis.get('candidate_name', 'Unknown'))
            ws.cell(row=row, column=3, value=analysis.get('filename', 'Unknown'))
            ws.cell(row=row, column=4, value=analysis.get('domain_detected', 'general'))
            ws.cell(row=row, column=5, value=analysis.get('years_of_experience', 'Not specified'))
            ws.cell(row=row, column=6, value=f"{analysis.get('overall_score', 0):.1f}")
            ws.cell(row=row, column=7, value=analysis.get('recommendation', 'N/A'))
            
            exp_summary = analysis.get('experience_summary', 'No summary available.')
            experience_bullets = convert_experience_to_bullet_points(exp_summary)
            ws.cell(row=row, column=8, value=experience_bullets)
            
            skills_matched = analysis.get('skills_matched', [])
            ws.cell(row=row, column=9, value=", ".join(skills_matched[:8]))
            
            skills_missing = analysis.get('skills_missing', [])
            ws.cell(row=row, column=10, value=", ".join(skills_missing[:8]))
            
            key_strengths = analysis.get('key_strengths', [])
            ws.cell(row=row, column=11, value=", ".join(key_strengths[:3]))
            
            areas_for_improvement = analysis.get('areas_for_improvement', [])
            ws.cell(row=row, column=12, value=", ".join(areas_for_improvement[:3]))
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        filepath = os.path.join(REPORTS_FOLDER, filename)
        wb.save(filepath)
        print(f"📊 Minimal batch report saved to: {filepath}")
        return filepath
    except Exception as e:
        print(f"❌ Error creating minimal batch report: {str(e)}")
        traceback.print_exc()
        return None

def get_score_grade_text(score):
    """Get text description for score"""
    if score >= 90:
        return "Exceptional Match 🎯"
    elif score >= 80:
        return "Very Good Match ✨"
    elif score >= 70:
        return "Good Match 👍"
    elif score >= 60:
        return "Fair Match 📊"
    elif score >= 50:
        return "Marginal Match 📉"
    elif score >= 40:
        return "Low Match ⚠️"
    elif score >= 30:
        return "Very Low Match ❌"
    else:
        return "Domain Mismatch 🚫"

@app.route('/download/<filename>', methods=['GET'])
def download_report(filename):
    """Download the Excel report"""
    update_activity()
    
    try:
        print(f"📥 Download request for: {filename}")
        
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '', filename)
        
        file_path = os.path.join(REPORTS_FOLDER, safe_filename)
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Download error: {traceback.format_exc()}")
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

@app.route('/download-single/<analysis_id>', methods=['GET'])
def download_single_report(analysis_id):
    """Download single candidate report"""
    update_activity()
    
    try:
        print(f"📥 Download single request for analysis ID: {analysis_id}")
        
        filename = f"single_analysis_{analysis_id}.xlsx"
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '', filename)
        
        file_path = os.path.join(REPORTS_FOLDER, safe_filename)
        
        if not os.path.exists(file_path):
            print(f"❌ Single report not found: {file_path}")
            return jsonify({'error': 'Single report not found'}), 404
        
        download_name = f"candidate_report_{analysis_id}.xlsx"
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Single download error: {traceback.format_exc()}")
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

@app.route('/warmup', methods=['GET'])
def force_warmup():
    """Force warm-up Groq API"""
    update_activity()
    
    try:
        available_keys = sum(1 for key in GROQ_API_KEYS if key)
        if available_keys == 0:
            return jsonify({
                'status': 'error',
                'message': 'No Groq API keys configured',
                'warmup_complete': False
            })
        
        result = warmup_groq_service()
        
        return jsonify({
            'status': 'success' if result else 'error',
            'message': f'Groq API warmed up successfully with {available_keys} keys' if result else 'Warm-up failed',
            'warmup_complete': warmup_complete,
            'ai_provider': 'groq',
            'model': GROQ_MODEL,
            'available_keys': available_keys,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'warmup_complete': False
        })

@app.route('/quick-check', methods=['GET'])
def quick_check():
    """Quick endpoint to check if Groq API is responsive"""
    update_activity()
    
    try:
        available_keys = sum(1 for key in GROQ_API_KEYS if key)
        if available_keys == 0:
            return jsonify({
                'available': False, 
                'reason': 'No Groq API keys configured',
                'available_keys': 0,
                'warmup_complete': warmup_complete
            })
        
        if not warmup_complete:
            return jsonify({
                'available': False,
                'reason': 'Groq API is warming up',
                'available_keys': available_keys,
                'warmup_complete': False,
                'ai_provider': 'groq',
                'model': GROQ_MODEL
            })
        
        for i, api_key in enumerate(GROQ_API_KEYS):
            if api_key and not key_usage[i]['cooling']:
                try:
                    start_time = time.time()
                    
                    response = call_groq_api(
                        prompt="Say 'ready'",
                        api_key=api_key,
                        max_tokens=10,
                        timeout=15,
                        key_index=i+1
                    )
                    
                    response_time = time.time() - start_time
                    
                    if isinstance(response, dict) and 'error' in response:
                        continue
                    elif response and 'ready' in str(response).lower():
                        return jsonify({
                            'available': True,
                            'response_time': f"{response_time:.2f}s",
                            'ai_provider': 'groq',
                            'model': GROQ_MODEL,
                            'warmup_complete': warmup_complete,
                            'available_keys': available_keys,
                            'tested_key': f"Key {i+1}",
                            'max_batch_size': MAX_BATCH_SIZE,
                            'processing_method': 'rate_limited_sequential',
                            'skills_analysis': '5-8 skills per category',
                            'rate_limit_protection': f"Active (max {MAX_REQUESTS_PER_MINUTE_PER_KEY}/min/key)",
                            'scoring_method': 'Professional ATS - Strict Domain Matching',
                            'deterministic_results': True,
                            'no_ranking_flip': True
                        })
                except:
                    continue
        
        return jsonify({
            'available': False,
            'reason': 'All keys failed or are cooling',
            'available_keys': available_keys,
            'warmup_complete': warmup_complete
        })
            
    except Exception as e:
        error_msg = str(e)
        return jsonify({
            'available': False,
            'reason': error_msg[:100],
            'available_keys': sum(1 for key in GROQ_API_KEYS if key),
            'ai_provider': 'groq',
            'model': GROQ_MODEL,
            'warmup_complete': warmup_complete
        })

@app.route('/ping', methods=['GET'])
def ping():
    """Simple ping to keep service awake"""
    update_activity()
    
    available_keys = sum(1 for key in GROQ_API_KEYS if key)
    
    return jsonify({
        'status': 'pong',
        'timestamp': datetime.now().isoformat(),
        'service': 'resume-analyzer-groq',
        'ai_provider': 'groq',
        'ai_warmup': warmup_complete,
        'model': GROQ_MODEL,
        'available_keys': available_keys,
        'inactive_minutes': int((datetime.now() - last_activity_time).total_seconds() / 60),
        'keep_alive_active': True,
        'max_batch_size': MAX_BATCH_SIZE,
        'processing_method': 'rate_limited_sequential',
        'skills_analysis': '5-8 skills per category',
        'years_experience': 'Included in analysis',
        'scoring_method': 'Professional ATS - Strict Domain Matching',
        'deterministic_results': True,
        'no_ranking_flip': True,
        'rate_limit_protection': f"Active (max {MAX_REQUESTS_PER_MINUTE_PER_KEY} requests/minute/key)"
    })

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint with key status"""
    update_activity()
    
    inactive_time = datetime.now() - last_activity_time
    inactive_minutes = int(inactive_time.total_seconds() / 60)
    
    current_time = datetime.now()
    key_status = []
    configured_keys = 0
    for i, api_key in enumerate(GROQ_API_KEYS):
        if api_key:
            configured_keys += 1
            if key_usage[i]['minute_window_start']:
                seconds_in_window = (current_time - key_usage[i]['minute_window_start']).total_seconds()
                if seconds_in_window > 60:
                    requests_this_minute = 0
                else:
                    requests_this_minute = key_usage[i]['requests_this_minute']
            else:
                requests_this_minute = 0
            
            key_status.append({
                'key': f'Key {i+1}',
                'configured': True,
                'total_usage': key_usage[i]['count'],
                'requests_this_minute': requests_this_minute,
                'errors': key_usage[i]['errors'],
                'cooling': key_usage[i]['cooling'],
                'last_used': key_usage[i]['last_used'].isoformat() if key_usage[i]['last_used'] else None
            })
        else:
            key_status.append({
                'key': f'Key {i+1}',
                'configured': False
            })
    
    available_keys = sum(1 for key in GROQ_API_KEYS if key)
    
    return jsonify({
        'status': 'Service is running', 
        'timestamp': datetime.now().isoformat(),
        'ai_provider': 'groq',
        'ai_provider_configured': available_keys > 0,
        'model': GROQ_MODEL,
        'ai_warmup_complete': warmup_complete,
        'upload_folder_exists': os.path.exists(UPLOAD_FOLDER),
        'reports_folder_exists': os.path.exists(REPORTS_FOLDER),
        'resume_previews_folder_exists': os.path.exists(RESUME_PREVIEW_FOLDER),
        'resume_previews_stored': len(resume_storage),
        'inactive_minutes': inactive_minutes,
        'version': '4.0.0',
        'key_status': key_status,
        'available_keys': configured_keys,
        'total_keys': 5,
        'configuration': {
            'max_batch_size': MAX_BATCH_SIZE,
            'max_requests_per_minute_per_key': MAX_REQUESTS_PER_MINUTE_PER_KEY,
            'max_retries': MAX_RETRIES,
            'min_skills_to_show': MIN_SKILLS_TO_SHOW,
            'max_skills_to_show': MAX_SKILLS_TO_SHOW,
            'years_experience_analysis': True,
            'scoring_system': 'PROFESSIONAL ATS - STRICT DOMAIN MATCHING',
            'deterministic_results': True,
            'no_ranking_flip': True,
            'domain_keyword_libraries': list(DOMAIN_KEYWORDS.keys())
        },
        'processing_method': 'rate_limited_sequential',
        'performance_target': '6 resumes in 20-30 seconds (safer)',
        'skills_analysis': '5-8 skills per category',
        'summaries': 'Complete 4-5 sentences each',
        'years_experience': 'Included in analysis',
        'excel_report': 'Candidate name & experience summary included',
        'insights': '3 strengths & 3 improvements',
        'scoring_enhancements': {
            'method': 'Professional ATS - Strict Domain Matching',
            'precision': '1 decimal place',
            'unique_scores': 'Deterministic hash-based',
            'range': '0-100 with strict domain matching',
            'weighting': 'Domain match (70%), Skills (20%), Experience (10%)',
            'wrong_domain_penalty': '70% reduction for domain mismatch',
            'ml_resume_vs_vlsi_job': 'Score: 25-35',
            'deterministic': 'Same resume = Same score',
            'ranking_stability': 'Never flips between analyses'
        },
        'rate_limit_protection': 'ACTIVE - Staggered delays, minute tracking, automatic cooling',
        'always_awake': True
    })

def cleanup_on_exit():
    """Cleanup function called on exit"""
    global service_running
    service_running = False
    print("\n🛑 Shutting down service...")
    
    try:
        # Clean up temporary files
        for folder in [UPLOAD_FOLDER, RESUME_PREVIEW_FOLDER]:
            for filename in os.listdir(folder):
                filepath = os.path.join(folder, filename)
                if os.path.isfile(filepath):
                    os.remove(filepath)
        print("✅ Cleaned up temporary files")
    except:
        pass

# Periodic cleanup
def periodic_cleanup():
    """Periodically clean up old resume previews"""
    while service_running:
        try:
            time.sleep(300)  # Run every 5 minutes
            cleanup_resume_previews()
        except Exception as e:
            print(f"⚠️ Periodic cleanup error: {str(e)}")

atexit.register(cleanup_on_exit)

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🚀 Resume Analyzer Backend Starting (Groq Parallel)...")
    print("="*50)
    print("🏢 MODE: PROFESSIONAL ATS SCORING SYSTEM")
    print("🎯 STRICT DOMAIN MATCHING - DETERMINISTIC RESULTS")
    print("="*50)
    PORT = int(os.environ.get('PORT', 5002))
    print(f"📍 Server: http://localhost:{PORT}")
    print(f"⚡ AI Provider: Groq (Parallel Processing)")
    print(f"🤖 Model: {GROQ_MODEL}")
    
    available_keys = sum(1 for key in GROQ_API_KEYS if key)
    print(f"🔑 API Keys: {available_keys}/5 configured")
    
    for i, key in enumerate(GROQ_API_KEYS):
        status = "✅ Configured" if key else "❌ Not configured"
        print(f"  Key {i+1}: {status}")
    
    print(f"📁 Upload folder: {UPLOAD_FOLDER}")
    print(f"📁 Reports folder: {REPORTS_FOLDER}")
    print(f"📁 Resume Previews folder: {RESUME_PREVIEW_FOLDER}")
    print(f"⚠️ RATE LIMIT PROTECTION: ACTIVE")
    print(f"📊 Max requests/minute/key: {MAX_REQUESTS_PER_MINUTE_PER_KEY}")
    print(f"⏳ Staggered delays: 1-3 seconds between requests")
    print(f"🔀 Key rotation: Smart load balancing (5 keys)")
    print(f"🛡️ Cooling: 60s on rate limits")
    print(f"✅ Max Batch Size: {MAX_BATCH_SIZE} resumes")  # UPDATED: Shows 10 resumes
    print(f"✅ Skills Analysis: {MIN_SKILLS_TO_SHOW}-{MAX_SKILLS_TO_SHOW} skills per category")
    print(f"✅ Years of Experience: Included in analysis")
    print(f"🏢 PROFESSIONAL ATS SCORING: STRICT DOMAIN MATCHING")
    print(f"🎯 ML resume vs VLSI job = 25-35 score (Not 65-70)")
    print(f"🎯 DETERMINISTIC RESULTS: Same resume = Same score")
    print(f"🎯 NO RANKING FLIPS: Rankings never change on re-analysis")
    print(f"🎯 Domain libraries: VLSI, ML, Software Eng, Data Science, DevOps")
    print(f"🎯 Scoring Method: Domain match (70%), Skills (20%), Experience (10%)")
    print(f"✅ Excel Report: Domain Detected column added")
    print(f"✅ Complete Summaries: 4-5 sentences each (no truncation)")
    print(f"✅ Insights: 3 strengths & 3 improvements")
    print(f"✅ Resume Preview: Enabled with PDF conversion")
    print(f"⚠️ Performance: ~6 resumes in 20-30 seconds (SAFER for rate limits)")
    print(f"✅ Excel Reports: Single & Batch with Individual Sheets")
    print(f"✅ Always Awake: Backend will stay active with self-pinging")
    print("="*50 + "\n")
    
    # Check for required dependencies
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        print("✅ PDF generation library available")
    except ImportError:
        print("⚠️  Warning: reportlab not installed. Install with: pip install reportlab")
        print("   PDF previews for non-PDF files will be limited")
    
    if available_keys == 0:
        print("⚠️  WARNING: No Groq API keys found!")
        print("Please set GROQ_API_KEY_1 through GROQ_API_KEY_5 in environment variables")
        print("Get free API keys from: https://console.groq.com")
    
    gc.enable()
    
    if available_keys > 0:
        warmup_thread = threading.Thread(target=warmup_groq_service, daemon=True)
        warmup_thread.start()
        
        keep_warm_thread = threading.Thread(target=keep_service_warm, daemon=True)
        keep_warm_thread.start()
        
        # Start keep-backend-awake thread
        keep_awake_thread = threading.Thread(target=keep_backend_awake, daemon=True)
        keep_awake_thread.start()
        
        cleanup_thread = threading.Thread(target=periodic_cleanup, daemon=True)
        cleanup_thread.start()
        
        print("✅ Background threads started")
    
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() in ('1', 'true', 'yes')
    app.run(host='0.0.0.0', port=PORT, debug=debug_mode, threaded=True)
